"""
sdtm_spec_builder.py - Build a draft SDTM specification from reviewed aCRF metadata.

Handles all SDTM domain classes:
  - Demographics (DM)
  - Events (AE, DS, MH, DV, CE)
  - Interventions (CM, EX, EC, PR, SU)
  - Findings (VS, LB, EG, PE, QS, SC, DA, MB, MS, PC, PP)
  - Findings About Events (TU, TR, RS)
  - SUPP-- (any supplemental qualifier domain)

Usage:
  python sdtm_spec_builder.py acrf_metadata.xlsx --output sdtm_spec_draft.xlsx
  python sdtm_spec_builder.py acrf_metadata.xlsx --output sdtm_spec_draft.xlsx --offline
"""

import argparse
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── CDISC structural variables by domain class ──────────────────────

COMMON_STRUCTURAL = [
    ("STUDYID",    "Study Identifier",                       "Char",  20, "Assigned", "Req"),
    ("{D}SEQ",     "Sequence Number",                        "Num",    8, "Derived",  "Req"),
    ("USUBJID",    "Unique Subject Identifier",              "Char",  40, "Derived",  "Req"),
    ("DOMAIN",     "Domain Abbreviation",                    "Char",   2, "Assigned", "Req"),
]

DM_EXTRA = [
    ("RFSTDTC",   "Subject Reference Start Date/Time",      "Char",  19, "CRF",      "Exp"),
    ("RFENDTC",   "Subject Reference End Date/Time",        "Char",  19, "CRF",      "Exp"),
    ("RFXSTDTC",  "Date/Time of First Study Treatment",     "Char",  19, "CRF",      "Exp"),
    ("RFXENDTC",  "Date/Time of Last Study Treatment",      "Char",  19, "CRF",      "Exp"),
    ("SITEID",    "Study Site Identifier",                   "Char",  10, "CRF",      "Req"),
    ("INVID",     "Investigator Identifier",                 "Char",  10, "Assigned", "Perm"),
    ("INVNAM",    "Investigator Name",                       "Char",  60, "Assigned", "Perm"),
    ("COUNTRY",   "Country",                                "Char",   3, "Assigned", "Req"),
    ("ARMCD",     "Planned Arm Code",                       "Char",  20, "Assigned", "Req"),
    ("ARM",       "Description of Planned Arm",             "Char", 200, "CRF",      "Req"),
    ("ACTARMCD",  "Actual Arm Code",                        "Char",  20, "Derived",  "Req"),
    ("ACTARM",    "Description of Actual Arm",              "Char", 200, "Derived",  "Req"),
]

# Events class: AE, DS, MH, DV, CE
EVENTS_EXTRA = [
    ("{D}TERM",   "Reported Term for the Event",            "Char", 200, "CRF",      "Req"),
    ("{D}DECOD",  "Dictionary-Derived Term",                "Char", 200, "Derived",  "Perm"),
    ("{D}CAT",    "Category for Event",                     "Char",  40, "Assigned", "Perm"),
    ("{D}SCAT",   "Subcategory for Event",                  "Char",  40, "Assigned", "Perm"),
    ("{D}BODSYS", "Body System or Organ Class",             "Char", 200, "Derived",  "Perm"),
    ("{D}STDTC",  "Start Date/Time of Event",               "Char",  19, "CRF",      "Exp"),
    ("{D}ENDTC",  "End Date/Time of Event",                 "Char",  19, "CRF",      "Perm"),
    ("{D}STDY",   "Study Day of Start of Event",            "Num",    8, "Derived",  "Perm"),
    ("{D}ENDY",   "Study Day of End of Event",              "Num",    8, "Derived",  "Perm"),
    ("EPOCH",     "Epoch",                                  "Char",  40, "Derived",  "Perm"),
]

# Interventions class: CM, EX, EC, PR, SU
INTERVENTIONS_EXTRA = [
    ("{D}TRT",    "Reported Name of Treatment",             "Char", 200, "CRF",      "Req"),
    ("{D}DECOD",  "Standardized Treatment Name",            "Char", 200, "Derived",  "Perm"),
    ("{D}CAT",    "Category for Intervention",              "Char",  40, "Assigned", "Perm"),
    ("{D}DOSE",   "Dose per Administration",                "Num",    8, "CRF",      "Exp"),
    ("{D}DOSU",   "Dose Units",                             "Char",  40, "CRF",      "Exp"),
    ("{D}DOSFRQ", "Dosing Frequency per Interval",          "Char",  40, "CRF",      "Exp"),
    ("{D}ROUTE",  "Route of Administration",                "Char",  40, "CRF",      "Exp"),
    ("{D}STDTC",  "Start Date/Time of Intervention",        "Char",  19, "CRF",      "Exp"),
    ("{D}ENDTC",  "End Date/Time of Intervention",          "Char",  19, "CRF",      "Perm"),
    ("{D}STDY",   "Study Day of Start of Intervention",     "Num",    8, "Derived",  "Perm"),
    ("{D}ENDY",   "Study Day of End of Intervention",       "Num",    8, "Derived",  "Perm"),
    ("EPOCH",     "Epoch",                                  "Char",  40, "Derived",  "Perm"),
]

# Findings class: VS, LB, EG, PE, QS, SC, etc.
FINDINGS_EXTRA = [
    ("{D}TESTCD", "Short Name of Measurement",              "Char",   8, "Assigned", "Req"),
    ("{D}TEST",   "Name of Measurement",                    "Char",  40, "Assigned", "Req"),
    ("{D}CAT",    "Category for Findings",                  "Char",  40, "Assigned", "Perm"),
    ("{D}ORRES",  "Result or Finding in Original Units",    "Char", 200, "CRF",      "Exp"),
    ("{D}ORRESU", "Original Units",                         "Char",  40, "Assigned", "Exp"),
    ("{D}STRESC", "Character Result in Std Format",         "Char", 200, "Derived",  "Exp"),
    ("{D}STRESN", "Numeric Result in Standard Units",       "Num",    8, "Derived",  "Exp"),
    ("{D}STRESU", "Standard Units",                         "Char",  40, "Assigned", "Exp"),
    ("{D}STAT",   "Completion Status",                      "Char",   8, "Assigned", "Perm"),
    ("{D}REASND", "Reason Not Performed",                   "Char", 200, "CRF",      "Perm"),
]

# Findings About Events class: TU, TR, RS
FINDINGS_ABOUT_EVENTS_EXTRA = [
    ("{D}TESTCD", "Short Name of Assessment",               "Char",   8, "Assigned", "Req"),
    ("{D}TEST",   "Name of Assessment",                     "Char",  40, "Assigned", "Req"),
    ("{D}CAT",    "Category for Assessment",                "Char",  40, "Assigned", "Perm"),
    ("{D}ORRES",  "Result or Finding in Original Units",    "Char", 200, "CRF",      "Exp"),
    ("{D}STRESC", "Character Result in Std Format",         "Char", 200, "Derived",  "Exp"),
    ("{D}STRESN", "Numeric Result in Standard Units",       "Num",    8, "Derived",  "Exp"),
    ("{D}STRESU", "Standard Units",                         "Char",  40, "Assigned", "Exp"),
    ("{D}EVAL",   "Evaluator",                              "Char",  40, "CRF",      "Exp"),
    ("{D}LNKID",  "Link ID",                                "Char",  40, "Assigned", "Perm"),
]

TIMING_VARS = [
    ("VISITNUM",  "Visit Number",                           "Num",    8, "Derived",  "Exp"),
    ("VISIT",     "Visit Name",                             "Char",  40, "CRF",      "Exp"),
    ("{D}DTC",    "Date/Time of Collection",                "Char",  19, "CRF",      "Exp"),
    ("{D}DY",     "Study Day of Collection",                "Num",    8, "Derived",  "Perm"),
]

SUPP_STRUCTURAL = [
    ("STUDYID",   "Study Identifier",                       "Char",  20, "Assigned", "Req"),
    ("RDOMAIN",   "Related Domain Abbreviation",            "Char",   2, "Assigned", "Req"),
    ("USUBJID",   "Unique Subject Identifier",              "Char",  40, "Derived",  "Req"),
    ("IDVAR",     "Identifying Variable",                   "Char",   8, "Assigned", "Req"),
    ("IDVARVAL",  "Identifying Variable Value",             "Char",  40, "Derived",  "Req"),
    ("QNAM",      "Qualifier Variable Name",                "Char",   8, "Assigned", "Req"),
    ("QLABEL",    "Qualifier Variable Label",               "Char",  40, "Assigned", "Req"),
    ("QVAL",      "Data Value",                             "Char", 200, "CRF",      "Req"),
    ("QORIG",     "Origin",                                 "Char",  20, "Assigned", "Req"),
    ("QEVAL",     "Evaluator",                              "Char",  20, "Assigned", "Perm"),
]

# Domain classification
EVENTS_DOMAINS = {"AE", "DS", "MH", "DV", "CE"}
INTERVENTIONS_DOMAINS = {"CM", "EC", "EX", "PR", "SU"}
FINDINGS_DOMAINS = {"VS", "LB", "EG", "PE", "QS", "SC", "DA", "MB", "MS", "PC", "PP"}
FINDINGS_ABOUT_EVENTS_DOMAINS = {"TU", "TR", "RS"}


# ── Read reviewed aCRF metadata ─────────────────────────────────────

def read_acrf_metadata(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["By Domain"]
    headers = [cell.value for cell in ws[1]]
    domains = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        status = (record.get("Review Status") or "").strip().upper()
        if status == "DELETE":
            continue
        domain = record.get("Domain", "")
        if domain:
            domains.setdefault(domain, []).append(record)
    wb.close()
    return domains


# ── Build structural variables per domain ───────────────────────────

def _expand(var_tuple, domain):
    name, label, vtype, length, origin, core = var_tuple
    name = name.replace("{D}", domain)
    label = label.replace("{D}", domain)
    return {
        "variable": name, "label": label, "type": vtype, "length": length,
        "origin": origin, "core": core, "codelist": "", "crf_page": "",
        "derivation": "", "source": "structural",
    }


def _get_domain_class(domain):
    if domain.startswith("SUPP"):
        return "SUPP"
    if domain == "DM":
        return "DM"
    if domain in EVENTS_DOMAINS:
        return "Events"
    if domain in INTERVENTIONS_DOMAINS:
        return "Interventions"
    if domain in FINDINGS_DOMAINS:
        return "Findings"
    if domain in FINDINGS_ABOUT_EVENTS_DOMAINS:
        return "Findings About Events"
    return "General"


def get_structural_vars(domain):
    dclass = _get_domain_class(domain)

    if dclass == "SUPP":
        return [_expand(v, domain) for v in SUPP_STRUCTURAL]

    result = [_expand(v, domain) for v in COMMON_STRUCTURAL]

    if dclass == "DM":
        result += [_expand(v, domain) for v in DM_EXTRA]
    elif dclass == "Events":
        result += [_expand(v, domain) for v in EVENTS_EXTRA]
    elif dclass == "Interventions":
        result += [_expand(v, domain) for v in INTERVENTIONS_EXTRA]
    elif dclass == "Findings":
        result += [_expand(v, domain) for v in FINDINGS_EXTRA]
        result += [_expand(v, domain) for v in TIMING_VARS]
    elif dclass == "Findings About Events":
        result += [_expand(v, domain) for v in FINDINGS_ABOUT_EVENTS_EXTRA]
        result += [_expand(v, domain) for v in TIMING_VARS]
    else:
        result += [_expand(v, domain) for v in TIMING_VARS]

    return result


# ── Claude API enrichment ───────────────────────────────────────────

def _call_claude(prompt):
    import anthropic
    client = anthropic.Anthropic()
    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )
    return response.content[0].text


def enrich_with_claude(domain, crf_fields):
    is_supp = domain.startswith("SUPP")
    dclass = _get_domain_class(domain)

    field_lines = []
    for f in crf_fields:
        line = f"  {f['Variable']}: CRF label = \"{f['CRF Label'] or ''}\" | codelist = \"{f['Codelist'] or ''}\" | qualifier = \"{f['Qualifier'] or ''}\""
        field_lines.append(line)

    if is_supp:
        parent = domain.replace("SUPP", "")
        prompt = f"""You are an expert CDISC SDTM programmer. These supplemental qualifier variables were extracted from an aCRF for {domain} (parent: {parent}):

{chr(10).join(field_lines)}

For each, provide QNAM metadata. Return ONLY a JSON array where each element has:
  "variable": QNAM value, "label": QLABEL value, "type": "Char", "length": 200,
  "origin": "CRF" or "Derived", "core": "Perm",
  "codelist": codelist name or "", "derivation": ""

Return valid JSON only, no markdown fences."""
    else:
        prompt = f"""You are an expert CDISC SDTM programmer. Variables extracted from an aCRF for domain {domain} (class: {dclass}):

{chr(10).join(field_lines)}

Provide correct SDTM metadata. Return ONLY a JSON array where each element has:
  "variable": SDTM variable name, "label": proper CDISC variable label,
  "type": "Char" or "Num", "length": integer, "origin": "CRF"/"Derived"/"Assigned",
  "core": "Req"/"Exp"/"Perm", "codelist": codelist name or "", "derivation": note or ""

Rules: use standard CDISC labels (not CRF labels). For findings with qualifiers like VSTESTCD=SYSBP, produce ONE row per unique variable. Only CRF-sourced variables, not structural.

Return valid JSON only, no markdown fences."""

    try:
        raw = _call_claude(prompt)
        cleaned = raw.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("\n", 1)[1]
        if cleaned.endswith("```"):
            cleaned = cleaned.rsplit("```", 1)[0]
        return json.loads(cleaned.strip())
    except Exception as e:
        print(f"  Warning: Claude API failed for {domain}: {e}")
        return _fallback_enrich(domain, crf_fields)


def _fallback_enrich(domain, crf_fields):
    is_supp = domain.startswith("SUPP")
    results = []
    seen = set()
    for f in crf_fields:
        var = f["Variable"]
        if var in seen:
            continue
        seen.add(var)
        if is_supp:
            vtype, length = "Char", 200
            core = "Perm"
        elif var.endswith("DTC"):
            vtype, length, core = "Char", 19, "Exp"
        elif var.endswith("STRESN") or var.endswith("SEQ") or var == "AGE" or var.endswith("DOSE"):
            vtype, length, core = "Num", 8, "Exp"
        elif var.endswith("CD") or var.endswith("FL") or var in ("SEX", "DOMAIN"):
            vtype, length, core = "Char", 8, "Exp"
        elif var.endswith("TERM") or var.endswith("TRT") or var.endswith("DECOD") or var.endswith("BODSYS"):
            vtype, length, core = "Char", 200, "Exp"
        else:
            vtype, length, core = "Char", 40, "Exp"
        results.append({
            "variable": var,
            "label": (f["CRF Label"] or var).rstrip(":").strip(),
            "type": vtype, "length": length, "origin": "CRF",
            "core": core, "codelist": f.get("Codelist") or "", "derivation": "",
        })
    return results


# ── Merge structural + CRF-enriched variables ──────────────────────

def build_domain_spec(domain, crf_fields, use_api=True):
    is_supp = domain.startswith("SUPP")
    dclass = _get_domain_class(domain)
    tag = f" [{dclass}]"
    print(f"\n  Building: {domain}{tag} ({len(crf_fields)} CRF fields)")

    structural = get_structural_vars(domain)
    structural_names = {v["variable"] for v in structural}

    enriched = enrich_with_claude(domain, crf_fields) if use_api else _fallback_enrich(domain, crf_fields)

    crf_vars = []
    seen = set()
    for e in enriched:
        var = e["variable"]
        if var in seen:
            continue
        seen.add(var)
        crf_page = ""
        for f in crf_fields:
            if f["Variable"] == var:
                crf_page = f.get("Page Title") or ""
                break
        if var in structural_names:
            for sv in structural:
                if sv["variable"] == var:
                    sv["codelist"] = e.get("codelist", "")
                    sv["crf_page"] = crf_page
                    if e.get("origin") == "CRF":
                        sv["origin"] = "CRF"
                    break
            continue
        crf_vars.append({
            "variable": var, "label": e.get("label", var),
            "type": e.get("type", "Char"), "length": e.get("length", 40),
            "origin": e.get("origin", "CRF"), "core": e.get("core", "Exp"),
            "codelist": e.get("codelist", ""), "crf_page": crf_page,
            "derivation": e.get("derivation", ""), "source": "crf",
        })

    all_vars = structural + crf_vars
    n_s = len(structural)
    n_c = len(crf_vars)

    if is_supp:
        print(f"    {n_s} structural (SUPP) + {n_c} qualifier values = {len(all_vars)}")
        print(f"    QNAM values: {', '.join(v['variable'] for v in crf_vars)}")
    else:
        print(f"    {n_s} structural + {n_c} CRF = {len(all_vars)} total")
    return all_vars


# ── Write the draft SDTM spec to Excel ──────────────────────────────

def write_sdtm_spec(domain_specs, output_path):
    wb = openpyxl.Workbook()

    hfont = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bfont = Font(name="Arial", size=10)
    struct_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    crf_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    supp_s_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    supp_c_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
    tborder = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["Variable", "Label", "Type", "Length", "Origin",
               "Core", "Codelist", "CRF Page", "Derivation"]
    col_w = [14, 45, 8, 8, 10, 8, 20, 20, 40]

    # Cover sheet
    cover = wb.active
    cover.title = "Cover"
    cover.cell(row=1, column=1, value="SDTM Specification Draft").font = Font(
        name="Arial", bold=True, size=16, color="1A3C6E")
    cover.cell(row=2, column=1, value="Generated by SpecGen - Phase 5b").font = Font(
        name="Arial", size=11, color="666666")
    cover.cell(row=3, column=1, value="Review each domain sheet, then pass to Phase 5c.").font = Font(
        name="Arial", size=10)
    cover.cell(row=4, column=1, value="Yellow=structural | Blue=CRF | Green=SUPP").font = Font(
        name="Arial", size=9, color="666666")

    for ci, h in enumerate(["Domain", "Class", "Variables", "Structural", "CRF"], 1):
        cell = cover.cell(row=6, column=ci, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.border = tborder
    for ci, w in enumerate([12, 22, 12, 12, 12], 1):
        cover.column_dimensions[get_column_letter(ci)].width = w

    total = 0
    for ri, domain in enumerate(sorted(domain_specs), 7):
        vl = domain_specs[domain]
        ns = sum(1 for v in vl if v.get("source") == "structural")
        nc = len(vl) - ns
        dc = _get_domain_class(domain)
        cover.cell(row=ri, column=1, value=domain).font = bfont
        cover.cell(row=ri, column=1).border = tborder
        cover.cell(row=ri, column=2, value=dc).font = bfont
        cover.cell(row=ri, column=2).border = tborder
        cover.cell(row=ri, column=3, value=len(vl)).font = bfont
        cover.cell(row=ri, column=3).border = tborder
        cover.cell(row=ri, column=4, value=ns).font = bfont
        cover.cell(row=ri, column=4).border = tborder
        cover.cell(row=ri, column=5, value=nc).font = bfont
        cover.cell(row=ri, column=5).border = tborder
        if domain.startswith("SUPP"):
            for ci in range(1, 6):
                cover.cell(row=ri, column=ci).fill = supp_s_fill
        total += len(vl)

    tr = 7 + len(domain_specs)
    bold = Font(name="Arial", bold=True, size=10)
    cover.cell(row=tr, column=1, value="TOTAL").font = bold
    cover.cell(row=tr, column=1).border = tborder
    cover.cell(row=tr, column=3, value=total).font = bold
    cover.cell(row=tr, column=3).border = tborder

    # Domain sheets
    for domain in sorted(domain_specs):
        is_supp = domain.startswith("SUPP")
        vl = domain_specs[domain]
        ws = wb.create_sheet(title=domain)

        if is_supp:
            parent = domain.replace("SUPP", "")
            ws.cell(row=1, column=1,
                    value=f"SUPP-- Structure: Each CRF field becomes a QNAM row in {domain}").font = Font(
                name="Arial", bold=True, size=10, color="006600")
            ws.cell(row=2, column=1,
                    value=f"RDOMAIN={parent} | IDVAR={parent}SEQ | QORIG=CRF").font = Font(
                name="Arial", size=9, color="666666")
            sr = 4
        else:
            dc = _get_domain_class(domain)
            ws.cell(row=1, column=1, value=f"Domain: {domain} ({dc})").font = Font(
                name="Arial", bold=True, size=11, color="1A3C6E")
            sr = 3

        for ci, (h, w) in enumerate(zip(headers, col_w), 1):
            cell = ws.cell(row=sr, column=ci, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = halign
            cell.border = tborder
            ws.column_dimensions[get_column_letter(ci)].width = w

        for ri, v in enumerate(vl, sr + 1):
            vals = [v["variable"], v["label"], v["type"], v["length"],
                    v["origin"], v["core"], v["codelist"],
                    v.get("crf_page", ""), v.get("derivation", "")]
            if is_supp:
                fill = supp_s_fill if v.get("source") == "structural" else supp_c_fill
            else:
                fill = struct_fill if v.get("source") == "structural" else crf_fill
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = bfont
                cell.fill = fill
                cell.border = tborder

        ws.freeze_panes = f"A{sr + 1}"
        lr = sr + len(vl)
        ws.auto_filter.ref = f"A{sr}:{get_column_letter(len(headers))}{lr}"

    wb.save(output_path)

    ns = sum(1 for d in domain_specs if not d.startswith("SUPP"))
    np = sum(1 for d in domain_specs if d.startswith("SUPP"))
    print(f"\nDraft SDTM spec written to {output_path}")
    print(f"  {ns} standard + {np} SUPP domain sheets + Cover")
    print(f"  {total} total variables")


# ── Main pipeline ──────────────────────────────────────────────────

def build_sdtm_spec(metadata_path, output_path, use_api=True):
    print(f"Reading reviewed aCRF metadata: {metadata_path}")
    domain_fields = read_acrf_metadata(metadata_path)

    std = [d for d in sorted(domain_fields) if not d.startswith("SUPP")]
    supp = [d for d in sorted(domain_fields) if d.startswith("SUPP")]
    print(f"  Standard domains ({len(std)}): {', '.join(std)}")
    print(f"  SUPP domains ({len(supp)}): {', '.join(supp)}")

    specs = {}
    for d in std + supp:
        specs[d] = build_domain_spec(d, domain_fields[d], use_api=use_api)

    write_sdtm_spec(specs, output_path)
    return specs


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Build draft SDTM spec from reviewed aCRF metadata.")
    parser.add_argument("metadata", help="Path to reviewed acrf_metadata.xlsx")
    parser.add_argument("--output", "-o", default="sdtm_spec_draft.xlsx")
    parser.add_argument("--offline", action="store_true",
                        help="Skip Claude API, use fallback inference")
    args = parser.parse_args()
    build_sdtm_spec(args.metadata, args.output, use_api=not args.offline)
