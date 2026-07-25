"""
acrf_parser.py - Extract SDTM annotations from an annotated CRF (aCRF) PDF.

How aCRFs work:
  - Printed CRF pages show field labels in black (e.g. "Date of Birth:")
  - Blue text next to each field shows the SDTM mapping (e.g. DM.BRTHDTC)
  - Some fields also have a codelist hint (e.g. "Codelist: SEX (M, F)")
  - Non-standard fields get SUPP-- annotations (e.g. SUPPAE.AEACNOTH)

What this module does:
  1. Opens the PDF page by page
  2. Splits characters into blue (annotations) vs black (field labels)
  3. Groups characters into text lines
  4. Identifies DOMAIN.VARIABLE patterns in blue lines (including SUPP--)
  5. Matches each annotation to the nearest black field label on the same row
  6. Captures any codelist hint associated with the annotation
  7. Returns a list of dicts: one per annotated field

Usage:
  # Console summary
  python acrf_parser.py sample_acrf.pdf

  # Export to Excel for review
  python acrf_parser.py sample_acrf.pdf --output acrf_metadata.xlsx

  # As a module
  from acrf_parser import parse_acrf, export_to_excel
  fields = parse_acrf("sample_acrf.pdf")
  export_to_excel(fields, "acrf_metadata.xlsx")
"""

import re
import argparse
import pdfplumber


# ── Colour detection ────────────────────────────────────────────────

def _is_blue(char):
    """Return True if this character is rendered in blue (annotation colour)."""
    color = char.get("non_stroking_color")
    if not color or len(color) < 3:
        return False
    r, g, b = float(color[0]), float(color[1]), float(color[2])
    return r < 0.1 and g < 0.1 and b > 0.7


# ── Character grouping ─────────────────────────────────────────────

def _chars_to_lines(chars, y_tolerance=3, x_gap=8):
    """
    Group a list of pdfplumber char dicts into text lines.
    Characters on the same y (within y_tolerance) are merged left-to-right.
    A gap wider than x_gap pixels inserts a space.
    """
    if not chars:
        return []

    chars = sorted(chars, key=lambda c: (round(c["top"]), c["x0"]))

    lines = []
    current = {"text": chars[0]["text"], "top": chars[0]["top"],
               "x0": chars[0]["x0"], "x1": chars[0]["x1"]}

    for ch in chars[1:]:
        same_line = abs(ch["top"] - current["top"]) <= y_tolerance
        if same_line:
            gap = ch["x0"] - current["x1"]
            if gap > x_gap:
                current["text"] += " "
            current["text"] += ch["text"]
            current["x1"] = max(current["x1"], ch["x1"])
        else:
            lines.append(current)
            current = {"text": ch["text"], "top": ch["top"],
                       "x0": ch["x0"], "x1": ch["x1"]}

    lines.append(current)
    return lines


# ── Annotation parsing ──────────────────────────────────────────────

# Matches: DM.BRTHDTC, AE.AETERM, SUPPAE.AEACNOTH, SUPPDM.COMPLT
# Domain can be 2+ chars (2 for standard, 4-6 for SUPP--)
_DOMAIN_VAR_RE = re.compile(r"^([A-Z]{2,6})\.([A-Z0-9_]+)$")

# Matches "Codelist: SOMETHING"
_CODELIST_RE = re.compile(r"^Codelist:\s*(.+)$", re.IGNORECASE)

# Matches "VSTESTCD=SYSBP" style qualifiers
_QUALIFIER_RE = re.compile(r"^([A-Z]+)=([A-Z0-9]+)$")


def _extract_page_title(black_lines):
    """Grab the page title from the topmost meaningful black text."""
    for line in sorted(black_lines, key=lambda l: l["top"]):
        text = line["text"].strip()
        if "Study " in text or "annotation" in text.lower() or "CRF field" in text:
            continue
        if len(text) > 3:
            return text
    return "Unknown"


def _find_nearest_label(annotation_line, black_lines, y_tolerance=15):
    """Find the black text line closest to the annotation on the same row, to its left."""
    target_top = annotation_line["top"]
    candidates = []

    for bl in black_lines:
        if abs(bl["top"] - target_top) > y_tolerance:
            continue
        if bl["x0"] < annotation_line["x0"]:
            candidates.append(bl)

    if not candidates:
        for bl in black_lines:
            if abs(bl["top"] - target_top) > 30:
                continue
            if bl["x0"] < annotation_line["x0"]:
                candidates.append(bl)

    if not candidates:
        return None

    candidates.sort(key=lambda bl: annotation_line["x0"] - bl["x1"])
    return candidates[0]["text"].strip()


# ── Main parser ─────────────────────────────────────────────────────

def parse_acrf(pdf_path):
    """
    Parse an aCRF PDF and return structured SDTM metadata.

    Returns list of dicts with keys:
      page, domain, variable, crf_label, codelist, qualifier, page_title, is_supp
    """
    results = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            all_chars = page.chars
            if not all_chars:
                continue

            blue_chars = [c for c in all_chars if _is_blue(c)]
            black_chars = [c for c in all_chars if not _is_blue(c)]

            blue_lines = _chars_to_lines(blue_chars)
            black_lines = _chars_to_lines(black_chars)

            page_title = _extract_page_title(black_lines)

            # Build codelist lookup
            codelist_by_top = {}
            for bl in blue_lines:
                m = _CODELIST_RE.match(bl["text"].strip())
                if m:
                    codelist_by_top[bl["top"]] = m.group(1).strip()

            # Process each blue line
            for bl in blue_lines:
                text = bl["text"].strip()

                if "annotation" in text.lower() or "CRF field" in text:
                    continue
                if _CODELIST_RE.match(text):
                    continue

                m = _DOMAIN_VAR_RE.match(text)
                if not m:
                    continue

                domain = m.group(1)
                variable = m.group(2)
                is_supp = domain.startswith("SUPP")

                crf_label = _find_nearest_label(bl, black_lines)

                codelist = None
                qualifier = None
                for cl_top, cl_text in codelist_by_top.items():
                    if abs(cl_top - bl["top"]) < 20 and cl_top > bl["top"]:
                        qm = _QUALIFIER_RE.match(cl_text)
                        if qm:
                            qualifier = cl_text
                        else:
                            codelist = cl_text
                        break

                results.append({
                    "page": page_num,
                    "domain": domain,
                    "variable": variable,
                    "crf_label": crf_label,
                    "codelist": codelist,
                    "qualifier": qualifier,
                    "page_title": page_title,
                    "is_supp": is_supp,
                })

    return results


# ── Excel export ────────────────────────────────────────────────────

def export_to_excel(fields, output_path):
    """
    Write extracted aCRF metadata to an Excel workbook for human review.

    Creates two sheets:
      - "By Domain"  : one row per field, sorted by domain then variable
      - "Summary"    : domain-level counts
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: By Domain ──────────────────────────────────────────

    ws = wb.active
    ws.title = "By Domain"

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="Arial", size=10)
    std_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    supp_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "Domain", "Variable", "CRF Label", "Codelist",
        "Qualifier", "CRF Page", "Page Title", "SUPP?", "Review Status"
    ]
    col_widths = [12, 18, 35, 30, 20, 10, 22, 8, 15]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    sorted_fields = sorted(fields, key=lambda f: (f["domain"], f["variable"]))

    for row_idx, f in enumerate(sorted_fields, start=2):
        row_data = [
            f["domain"],
            f["variable"],
            f["crf_label"] or "",
            f["codelist"] or "",
            f["qualifier"] or "",
            f["page"],
            f["page_title"],
            "Yes" if f["is_supp"] else "",
            "",
        ]
        fill = supp_fill if f["is_supp"] else std_fill
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.border = thin_border
            cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(sorted_fields) + 1}"

    # ── Sheet 2: Summary ────────────────────────────────────────────

    ws2 = wb.create_sheet(title="Summary")

    for col_idx, header in enumerate(["Domain", "Variable Count", "Type", "CRF Page(s)"], start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 30

    domain_counts = {}
    domain_pages = {}
    for f in fields:
        d = f["domain"]
        domain_counts[d] = domain_counts.get(d, 0) + 1
        domain_pages.setdefault(d, set()).add(f["page_title"])

    for row_idx, domain in enumerate(sorted(domain_counts), start=2):
        dtype = "SUPP" if domain.startswith("SUPP") else "Standard"
        ws2.cell(row=row_idx, column=1, value=domain).font = body_font
        ws2.cell(row=row_idx, column=1).border = thin_border
        ws2.cell(row=row_idx, column=2, value=domain_counts[domain]).font = body_font
        ws2.cell(row=row_idx, column=2).border = thin_border
        ws2.cell(row=row_idx, column=3, value=dtype).font = body_font
        ws2.cell(row=row_idx, column=3).border = thin_border
        ws2.cell(row=row_idx, column=4, value=", ".join(sorted(domain_pages[domain]))).font = body_font
        ws2.cell(row=row_idx, column=4).border = thin_border

    total_row = len(domain_counts) + 2
    bold = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=total_row, column=1, value="TOTAL").font = bold
    ws2.cell(row=total_row, column=1).border = thin_border
    ws2.cell(row=total_row, column=2, value=len(fields)).font = bold
    ws2.cell(row=total_row, column=2).border = thin_border

    ws2.freeze_panes = "A2"

    wb.save(output_path)

    n_supp = sum(1 for f in fields if f["is_supp"])
    n_std = len(fields) - n_supp
    print(f"Exported {len(fields)} fields to {output_path}")
    print(f"  {n_std} standard + {n_supp} SUPP fields")
    print(f"  Green rows = SUPP domain fields")
    print(f"  Blue rows = standard domain fields")


# ── Console summary ────────────────────────────────────────────────

def print_summary(fields):
    """Print a readable summary grouped by domain."""
    domains = {}
    for f in fields:
        domains.setdefault(f["domain"], []).append(f)

    n_supp = sum(1 for f in fields if f["is_supp"])
    print(f"\n{'='*70}")
    print(f"aCRF Extraction Summary: {len(fields)} fields across "
          f"{len(domains)} domains ({n_supp} SUPP fields)")
    print(f"{'='*70}")

    for domain in sorted(domains):
        vars_list = domains[domain]
        tag = " [SUPP]" if domain.startswith("SUPP") else ""
        print(f"\n  Domain: {domain}{tag} ({len(vars_list)} variables)")
        print(f"  {'Variable':<15} {'CRF Label':<35} {'Codelist/Qualifier'}")
        print(f"  {'-'*15} {'-'*35} {'-'*25}")
        for v in vars_list:
            extra = v["codelist"] or v["qualifier"] or ""
            label = (v["crf_label"] or "(no label found)")[:35]
            print(f"  {v['variable']:<15} {label:<35} {extra}")


# ── CLI entry point ─────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Parse an annotated CRF (aCRF) PDF and extract SDTM metadata."
    )
    parser.add_argument("pdf", help="Path to the aCRF PDF file")
    parser.add_argument("--output", "-o",
                        help="Export to Excel (.xlsx) for review")

    args = parser.parse_args()

    print(f"Parsing aCRF: {args.pdf}")
    fields = parse_acrf(args.pdf)
    print_summary(fields)

    if args.output:
        print()
        export_to_excel(fields, args.output)
