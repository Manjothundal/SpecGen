"""
protocol_parser.py - Extract trial design metadata from a protocol PDF.

Produces structured data for trial design SDTM domains:
  TA - Trial Arms (arm codes, descriptions, element sequences)
  TE - Trial Elements (screening, treatment, follow-up with durations)
  TV - Trial Visits (visit numbers, names, target days, windows)
  TI - Trial Inclusion/Exclusion criteria
  TS - Trial Summary (sponsor, phase, indication, title, etc.)

Two modes:
  --offline : regex/pattern-based extraction (works without API)
  default   : sends extracted text to Claude API for structured parsing

Usage:
  python protocol_parser.py sample_protocol.pdf --output protocol_metadata.xlsx
  python protocol_parser.py sample_protocol.pdf --output protocol_metadata.xlsx --offline

  # As a module
  from protocol_parser import parse_protocol, export_to_excel
  data = parse_protocol("sample_protocol.pdf")
  export_to_excel(data, "protocol_metadata.xlsx")
"""

import re
import argparse
import json
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── PDF text extraction ─────────────────────────────────────────────

def extract_full_text(pdf_path):
    """Extract all text from the protocol PDF, page by page."""
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            pages.append(text)
    return pages


def extract_tables(pdf_path):
    """Extract tables from the PDF (pdfplumber table detection)."""
    all_tables = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            tables = page.extract_tables()
            for table in tables:
                if table:
                    all_tables.append({"page": page_num, "rows": table})
    return all_tables


# ── Offline pattern-based extraction ────────────────────────────────

def _extract_ts_offline(full_text):
    """Extract Trial Summary (TS) key-value pairs from synopsis section."""
    ts_records = []
    # Common TS parameter codes mapped to text labels
    ts_patterns = {
        "SPONSOR":    r"Sponsor[:\s]+(.+?)(?:\n|$)",
        "PROTID":     r"Protocol\s+Number[:\s]+(.+?)(?:\n|$)",
        "TITLE":      r"Protocol\s+Title[:\s]+(.+?)(?:\n|$)",
        "STITLE":     r"Short\s+Title[:\s]+(.+?)(?:\n|$)",
        "TPHASE":     r"Phase[:\s]+(Phase\s+\S+)",
        "INDIC":      r"Indication[:\s]+(.+?)(?:\n|$)",
        "STYPE":      r"Study\s+Type[:\s]+(.+?)(?:\n|$)",
        "PLESSION":   r"Number\s+of\s+Subjects[:\s]+(.+?)(?:\n|$)",
        "NSITE":      r"Number\s+of\s+Sites[:\s]+(.+?)(?:\n|$)",
        "TRDUR":      r"Treatment\s+Duration[:\s]+(.+?)(?:\n|$)",
        "RANDQT":     r"Randomization\s+Ratio[:\s]+(.+?)(?:\n|$)",
        "BLIND":      r"Blinding[:\s]+(.+?)(?:\n|$)",
        "COMPTRT":    r"Comparator[:\s]+(.+?)(?:\n|$)",
        "PRIENDPT":   r"Primary\s+Endpoint[:\s]+(.+?)(?:\n|$)",
        "STATMETH":   r"Statistical\s+Method[:\s]+(.+?)(?:\n|$)",
        "REGAGENCY":  r"Regulatory\s+Agency[:\s]+(.+?)(?:\n|$)",
    }

    for parmcd, pattern in ts_patterns.items():
        m = re.search(pattern, full_text, re.IGNORECASE)
        if m:
            value = m.group(1).strip()
            ts_records.append({
                "TSPARMCD": parmcd,
                "TSPARM": parmcd.replace("_", " ").title(),
                "TSVAL": value,
            })

    return ts_records


def _extract_arms_offline(full_text):
    """Extract Trial Arms (TA) from arm/treatment tables."""
    arms = []
    # Look for patterns like "ARM A", "ARM B", "ARM C" with descriptions
    arm_pattern = re.compile(
        r"(?:ARM\s+([A-Z]))[,:\s]+(.+?)(?=ARM\s+[A-Z]|$)",
        re.IGNORECASE | re.DOTALL
    )

    # Also try table-style: number, arm code, description
    table_pattern = re.compile(
        r"(\d+)\s+(ARM\s+[A-Z])\s+(.+?)(?:\n|$)",
        re.IGNORECASE
    )

    matches = table_pattern.findall(full_text)
    if matches:
        for order, armcd, desc in matches:
            arms.append({
                "ARMCD": armcd.strip().replace("ARM ", "ARM"),
                "ARM": desc.strip(),
                "TAETORD": int(order),
            })
    else:
        # Fallback: try the simpler pattern
        for i, m in enumerate(arm_pattern.finditer(full_text), 1):
            arms.append({
                "ARMCD": f"ARM{m.group(1)}",
                "ARM": m.group(2).strip().split("\n")[0],
                "TAETORD": i,
            })

    return arms


def _extract_elements_offline(full_text):
    """Extract Trial Elements (TE) from study design section."""
    elements = []
    # Look for table rows with: order, code, element name, epoch, duration
    elem_pattern = re.compile(
        r"(\d+)\s+([A-Z]+)\s+(.+?)\s+(SCREENING|LEAD-IN|TREATMENT|FOLLOW-UP|WASHOUT)\s+(.+?)\s+(.+?)(?:\n|$)",
        re.IGNORECASE
    )

    for m in elem_pattern.finditer(full_text):
        elements.append({
            "ETCD": m.group(2).strip(),
            "ELEMENT": m.group(3).strip(),
            "EPOCH": m.group(4).strip().upper(),
            "TEDUR": m.group(5).strip(),
            "TESTRL": m.group(6).strip(),
        })

    return elements


def _extract_visits_offline(full_text):
    """Extract Trial Visits (TV) from visit schedule."""
    visits = []
    # Pattern: visit_num, visit_name, target_day, window, epoch
    visit_pattern = re.compile(
        r"(\d+)\s+(.+?)\s+(-?\d+|)\s+(.+?)\s+(SCREENING|LEAD-IN|TREATMENT|FOLLOW-UP|)\s*(?:\n|$)",
        re.IGNORECASE
    )

    for m in visit_pattern.finditer(full_text):
        vnum = m.group(1).strip()
        vname = m.group(2).strip()
        vday = m.group(3).strip()
        window = m.group(4).strip()

        # Filter out noise
        if not vname or len(vname) < 3:
            continue
        if vname.lower().startswith(("visit#", "visit name", "target")):
            continue

        visits.append({
            "VISITNUM": int(vnum) if vnum else None,
            "VISIT": vname,
            "VISITDY": int(vday) if vday and vday.lstrip("-").isdigit() else None,
            "WINDOW": window,
        })

    return visits


def _extract_ie_offline(full_text):
    """Extract Inclusion/Exclusion criteria (TI)."""
    criteria = []

    # Inclusion: IN01, IN02, etc.
    in_pattern = re.compile(r"(IN\d{2})\.\s*(.+?)(?=IN\d{2}\.|EX\d{2}\.|$)", re.DOTALL)
    for m in in_pattern.finditer(full_text):
        criteria.append({
            "IETESTCD": m.group(1).strip(),
            "IETEST": m.group(2).strip().replace("\n", " "),
            "IECAT": "INCLUSION",
        })

    # Exclusion: EX01, EX02, etc.
    ex_pattern = re.compile(r"(EX\d{2})\.\s*(.+?)(?=EX\d{2}\.|IN\d{2}\.|$)", re.DOTALL)
    for m in ex_pattern.finditer(full_text):
        criteria.append({
            "IETESTCD": m.group(1).strip(),
            "IETEST": m.group(2).strip().replace("\n", " "),
            "IECAT": "EXCLUSION",
        })

    return criteria


# ── Claude API extraction ───────────────────────────────────────────

def _call_claude(prompt):
    import anthropic
    client = anthropic.Anthropic()
    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=8192,
        messages=[{"role": "user", "content": prompt}],
    )
    return response.content[0].text


def _parse_json_response(raw):
    cleaned = raw.strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.split("\n", 1)[1]
    if cleaned.endswith("```"):
        cleaned = cleaned.rsplit("```", 1)[0]
    return json.loads(cleaned.strip())


def extract_with_claude(full_text):
    """Send protocol text to Claude for structured extraction of all trial design domains."""

    prompt = f"""You are an expert CDISC SDTM programmer. Extract trial design metadata from this clinical protocol text. Return ONLY a JSON object (no markdown fences) with these keys:

"TS": array of {{"TSPARMCD": "...", "TSPARM": "...", "TSVAL": "..."}},
"TA": array of {{"ARMCD": "...", "ARM": "description", "TAETORD": number}},
"TE": array of {{"ETCD": "code", "ELEMENT": "name", "EPOCH": "...", "TEDUR": "duration string", "TESTRL": "timing"}},
"TV": array of {{"VISITNUM": number, "VISIT": "name", "VISITDY": number_or_null, "WINDOW": "window string"}},
"TI": array of {{"IETESTCD": "IN01/EX01", "IETEST": "full criterion text", "IECAT": "INCLUSION/EXCLUSION"}}

TS parameter codes to extract: SPONSOR, PROTID, TITLE, STITLE, TPHASE, INDIC, STYPE, PLESSION (planned subjects), NSITE, TRDUR, RANDQT, BLIND, COMPTRT, PRIENDPT, STATMETH, REGAGENCY

Protocol text:
{full_text[:15000]}

Return valid JSON only."""

    try:
        raw = _call_claude(prompt)
        return _parse_json_response(raw)
    except Exception as e:
        print(f"  Warning: Claude API failed: {e}")
        print(f"  Falling back to offline extraction")
        return None


# ── Main parser ─────────────────────────────────────────────────────

def parse_protocol(pdf_path, use_api=True):
    """
    Parse a protocol PDF and return trial design metadata.

    Returns dict with keys: TS, TA, TE, TV, TI
    """
    print(f"Extracting text from: {pdf_path}")
    pages = extract_full_text(pdf_path)
    full_text = "\n".join(pages)
    print(f"  {len(pages)} pages, {len(full_text)} characters extracted")

    if use_api:
        result = extract_with_claude(full_text)
        if result:
            return result

    # Offline extraction
    print("  Using offline pattern-based extraction")
    return {
        "TS": _extract_ts_offline(full_text),
        "TA": _extract_arms_offline(full_text),
        "TE": _extract_elements_offline(full_text),
        "TV": _extract_visits_offline(full_text),
        "TI": _extract_ie_offline(full_text),
    }


# ── Excel export ────────────────────────────────────────────────────

def export_to_excel(data, output_path):
    """Write protocol metadata to Excel with one sheet per domain."""
    wb = openpyxl.Workbook()

    hfont = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bfont = Font(name="Arial", size=10)
    fill1 = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    fill2 = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    tborder = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

    def write_sheet(ws, headers, rows, col_widths):
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = halign
            cell.border = tborder
            ws.column_dimensions[get_column_letter(ci)].width = w
        for ri, row in enumerate(rows, 2):
            fill = fill1 if ri % 2 == 0 else fill2
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = bfont
                cell.fill = fill
                cell.border = tborder
                if ci == len(row):  # wrap last column
                    cell.alignment = Alignment(wrap_text=True)
        ws.freeze_panes = "A2"

    # TS - Trial Summary
    ws = wb.active
    ws.title = "TS"
    ts_rows = [[r["TSPARMCD"], r["TSPARM"], r["TSVAL"]] for r in data.get("TS", [])]
    write_sheet(ws, ["TSPARMCD", "TSPARM", "TSVAL"], ts_rows, [15, 25, 80])

    # TA - Trial Arms
    ws2 = wb.create_sheet("TA")
    ta_rows = [[r["ARMCD"], r["ARM"], r.get("TAETORD", "")] for r in data.get("TA", [])]
    write_sheet(ws2, ["ARMCD", "ARM", "TAETORD"], ta_rows, [15, 60, 12])

    # TE - Trial Elements
    ws3 = wb.create_sheet("TE")
    te_rows = [[r["ETCD"], r["ELEMENT"], r.get("EPOCH", ""), r.get("TEDUR", ""), r.get("TESTRL", "")]
               for r in data.get("TE", [])]
    write_sheet(ws3, ["ETCD", "ELEMENT", "EPOCH", "TEDUR", "TESTRL"], te_rows, [10, 25, 18, 18, 25])

    # TV - Trial Visits
    ws4 = wb.create_sheet("TV")
    tv_rows = [[r["VISITNUM"], r["VISIT"], r.get("VISITDY", ""), r.get("WINDOW", "")]
               for r in data.get("TV", [])]
    write_sheet(ws4, ["VISITNUM", "VISIT", "VISITDY", "WINDOW"], tv_rows, [12, 30, 12, 20])

    # TI - Trial Inclusion/Exclusion
    ws5 = wb.create_sheet("TI")
    ti_rows = [[r["IETESTCD"], r["IECAT"], r["IETEST"]] for r in data.get("TI", [])]
    write_sheet(ws5, ["IETESTCD", "IECAT", "IETEST"], ti_rows, [12, 15, 80])

    # Cover sheet
    cover = wb.create_sheet("Cover", 0)  # insert at position 0
    cover.cell(row=1, column=1, value="Protocol Metadata - Trial Design Domains").font = Font(
        name="Arial", bold=True, size=16, color="1A3C6E")
    cover.cell(row=2, column=1, value="Generated by SpecGen - Phase 5d").font = Font(
        name="Arial", size=11, color="666666")
    cover.cell(row=3, column=1, value="Review each domain sheet before SDTM program generation.").font = Font(
        name="Arial", size=10)

    summary = [
        ("TS", "Trial Summary", len(data.get("TS", []))),
        ("TA", "Trial Arms", len(data.get("TA", []))),
        ("TE", "Trial Elements", len(data.get("TE", []))),
        ("TV", "Trial Visits", len(data.get("TV", []))),
        ("TI", "Inclusion/Exclusion", len(data.get("TI", []))),
    ]
    for ci, h in enumerate(["Domain", "Description", "Records"], 1):
        cell = cover.cell(row=5, column=ci, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.border = tborder
    cover.column_dimensions["A"].width = 10
    cover.column_dimensions["B"].width = 25
    cover.column_dimensions["C"].width = 10

    for ri, (dom, desc, count) in enumerate(summary, 6):
        cover.cell(row=ri, column=1, value=dom).font = bfont
        cover.cell(row=ri, column=1).border = tborder
        cover.cell(row=ri, column=2, value=desc).font = bfont
        cover.cell(row=ri, column=2).border = tborder
        cover.cell(row=ri, column=3, value=count).font = bfont
        cover.cell(row=ri, column=3).border = tborder

    wb.save(output_path)

    total = sum(len(data.get(k, [])) for k in ["TS", "TA", "TE", "TV", "TI"])
    print(f"\nExported to {output_path}")
    print(f"  5 domain sheets (TS, TA, TE, TV, TI) + Cover")
    print(f"  {total} total records")
    for dom, desc, count in summary:
        print(f"    {dom}: {count} records")


# ── Console summary ────────────────────────────────────────────────

def print_summary(data):
    print(f"\n{'='*70}")
    print("Protocol Extraction Summary")
    print(f"{'='*70}")

    ts = data.get("TS", [])
    if ts:
        print(f"\n  TS - Trial Summary ({len(ts)} parameters)")
        for r in ts:
            val = r["TSVAL"][:60] + "..." if len(r["TSVAL"]) > 60 else r["TSVAL"]
            print(f"    {r['TSPARMCD']:<12} {val}")

    ta = data.get("TA", [])
    if ta:
        print(f"\n  TA - Trial Arms ({len(ta)} arms)")
        for r in ta:
            print(f"    {r['ARMCD']:<10} {r['ARM']}")

    te = data.get("TE", [])
    if te:
        print(f"\n  TE - Trial Elements ({len(te)} elements)")
        for r in te:
            print(f"    {r['ETCD']:<8} {r['ELEMENT']:<25} {r.get('EPOCH', ''):<15} {r.get('TEDUR', '')}")

    tv = data.get("TV", [])
    if tv:
        print(f"\n  TV - Trial Visits ({len(tv)} visits)")
        for r in tv:
            day = r.get("VISITDY", "")
            day_str = str(day) if day is not None else ""
            print(f"    {str(r['VISITNUM']):<6} {r['VISIT']:<30} Day {day_str:<8} {r.get('WINDOW', '')}")

    ti = data.get("TI", [])
    if ti:
        n_in = sum(1 for r in ti if r["IECAT"] == "INCLUSION")
        n_ex = sum(1 for r in ti if r["IECAT"] == "EXCLUSION")
        print(f"\n  TI - Inclusion/Exclusion ({len(ti)} criteria: {n_in} inclusion, {n_ex} exclusion)")
        for r in ti:
            text = r["IETEST"][:70] + "..." if len(r["IETEST"]) > 70 else r["IETEST"]
            print(f"    {r['IETESTCD']:<6} [{r['IECAT'][:3]}] {text}")


# ── CLI entry point ─────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Parse a protocol PDF and extract trial design metadata (TA, TE, TV, TI, TS).")
    parser.add_argument("pdf", help="Path to the protocol PDF")
    parser.add_argument("--output", "-o",
                        help="Export to Excel (.xlsx)")
    parser.add_argument("--offline", action="store_true",
                        help="Skip Claude API, use pattern-based extraction")
    args = parser.parse_args()

    data = parse_protocol(args.pdf, use_api=not args.offline)
    print_summary(data)

    if args.output:
        export_to_excel(data, args.output)
