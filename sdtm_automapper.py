"""
sdtm_automapper.py - Propose SDTM mappings directly from raw clinical data.

Every other SDTM path in this app (sdtm_spec_builder.py, sdtm_assembler.py)
starts from a reviewed spec — a human already decided which SDTM variable
each field belongs to. This module is for when that spec doesn't exist yet:
point it at a folder of raw datasets (however a site/vendor exported them —
SAS7BDAT or CSV, EDC field names, no CDISC structure at all) and it proposes
source-to-target mappings for a human to review, the same "machine drafts,
human decides" shape as the aCRF path.

How it works:
  1. Read each raw dataset in the folder (pyreadstat for .sas7bdat, pandas
     for .csv) and pull variable-level metadata: name, type, label (SAS
     datasets only — CSV has none), format, and up to 10 sample values.
  2. Send ONE Claude call per dataset (not one for the whole folder, and not
     pre-clustered by domain — a raw dataset is already a natural grouping;
     see build_mapping_prompt) asking it to classify every variable in that
     dataset against the CDISC SDTM domain list and propose a mapping.
  3. Normalize each returned mapping's confidence into exactly High/Medium/
     Low (an LLM-supplied numeric score isn't reliable enough to threshold
     meaningfully; asking for one of three buckets directly is), and default
     anything malformed or missing to Low — never silently drop a variable
     the model didn't map cleanly, flag it instead.
  4. Write TWO sheets to one .xlsx:
       "Review"    — the human-facing view: Source Dataset, Source Variable,
                      Source Label, Sample Values, SDTM Domain, SDTM
                      Variable, Mapping Logic, Confidence, Review Status —
                      rows tinted by confidence (Excel's own standard
                      good/neutral/bad green/amber/red) so low-confidence
                      rows are impossible to miss.
       "By Domain" — the SAME mappings reshaped into the exact column shape
                      sdtm_spec_builder.py's read_acrf_metadata() already
                      expects (Domain, Variable, CRF Label, Codelist,
                      Qualifier, CRF Page, Page Title, SUPP?, Review
                      Status), so after review this workbook can be handed
                      straight to sdtm_spec_builder.py with no conversion
                      step and no changes to that file. Only rows with a
                      real (non-UNKNOWN) domain and target variable are
                      included — an unmapped row has nothing to build a
                      spec sheet from and belongs in a human's queue, not a
                      domain sheet sdtm_spec_builder.py would try to build
                      structural variables around.
     Review on "Review" is for a human reading top to bottom; the actual
     accept/reject/fix edits that matter for the next step happen on "By
     Domain" (delete a row to drop it, same Review Status convention
     acrf_parser.py's own "By Domain" sheet uses).

Usage:
  python sdtm_automapper.py raw_data/ --output sdtm_automap.xlsx
  python sdtm_automapper.py raw_data/ --output sdtm_automap.xlsx --offline

  # Next step, once sdtm_automap.xlsx's "By Domain" sheet has been reviewed:
  python sdtm_spec_builder.py sdtm_automap.xlsx --output sdtm_spec_draft.xlsx
  python sdtm_assembler.py sdtm_spec_draft.xlsx --output sdtm_programs/

  # As a module
  from sdtm_automapper import run_automapper
  rows = run_automapper("raw_data/", "sdtm_automap.xlsx", use_api=True)
"""

import argparse
import json
import os

import pandas as pd
import pyreadstat

from generator import generate_api, generate_local

# 12 standard SDTM domains this app's SDTM pipeline already supports
# (sdtm_assembler.py/sdtm_spec_builder.py), plus SUPP-- for any of them —
# same list, stated here rather than imported, to keep this module
# standalone (it doesn't depend on sdtm_spec_builder.py to run; it only
# needs to produce output that file can read).
SDTM_DOMAINS = ["DM", "AE", "VS", "CM", "DS", "MH", "DV", "EG", "EX", "TU", "TR", "RS"]

CONFIDENCE_LEVELS = ("High", "Medium", "Low")

SAMPLE_VALUES_SHOWN = 10


# ── Reading raw datasets ────────────────────────────────────────────

def read_dataset(path):
    """Read one raw dataset into (dataframe, labels dict, formats dict).
    labels/formats are empty for CSV — plain CSV has no metadata beyond
    the column names themselves."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".sas7bdat":
        df, meta = pyreadstat.read_sas7bdat(path)
        return df, dict(meta.column_names_to_labels), dict(meta.original_variable_types)
    elif ext == ".csv":
        return pd.read_csv(path), {}, {}
    else:
        raise ValueError(f"Unsupported file type {ext!r} (expected .sas7bdat or .csv)")


def extract_variable_metadata(df, labels, formats):
    """One dict per column: variable name, Char/Num, label, format, and up
    to SAMPLE_VALUES_SHOWN non-missing sample values (as strings, so a date
    stored as a SAS numeric and a date stored as CSV text both show up the
    same way to the model)."""
    result = []
    for col in df.columns:
        series = df[col]
        is_numeric = pd.api.types.is_numeric_dtype(series)
        samples = series.dropna().astype(str).head(SAMPLE_VALUES_SHOWN).tolist()
        result.append({
            "variable": col,
            "type": "Num" if is_numeric else "Char",
            "label": (labels.get(col) or "").strip(),
            "format": (formats.get(col) or "").strip(),
            "samples": samples,
        })
    return result


# ── Claude mapping ──────────────────────────────────────────────────

def build_mapping_prompt(dataset_name, var_meta, sdtmig_version=None):
    lines = []
    for v in var_meta:
        parts = [f'{v["variable"]} ({v["type"]})']
        if v["label"]:
            parts.append(f'label="{v["label"]}"')
        if v["format"]:
            parts.append(f'format={v["format"]}')
        samples = ", ".join(v["samples"]) if v["samples"] else "(no non-missing values)"
        parts.append(f"sample values: {samples}")
        lines.append("  " + " | ".join(parts))

    domain_list = ", ".join(SDTM_DOMAINS)

    return f"""You are a CDISC SDTM expert. Given this raw clinical dataset's variable
metadata, classify which SDTM domain each variable maps to and propose a
source-to-target mapping.

Raw dataset: {dataset_name}

Variables:
{chr(10).join(lines)}

SDTM domains to consider: {domain_list}. If a variable is non-standard for
its domain (doesn't fit a standard SDTM variable), propose the matching
SUPP-- domain instead (e.g. SUPPAE) as sdtm_domain, with a QNAM-style
sdtm_var. If a variable doesn't belong in SDTM at all (e.g. an internal
tracking field), use sdtm_domain "UNKNOWN".

For EACH variable above, return one JSON object with:
  "source_var": the raw variable name, exactly as given above
  "sdtm_domain": the 2-letter domain code, "SUPP"+parent, or "UNKNOWN"
  "sdtm_var": the target SDTM variable name (e.g. AETERM, VSORRES), or ""
    if sdtm_domain is "UNKNOWN"
  "qualifier": for a Findings-class TESTCD/ORRES pair (VS/LB/EG), the
    implied qualifier like "VSTESTCD=SYSBP" — "" if not applicable
  "mapping_logic": one sentence explaining the mapping, or why it's UNKNOWN
  "confidence": exactly "High" (>90% sure this mapping is correct),
    "Medium" (70-90%), or "Low" (<70%, including anything ambiguous or
    UNKNOWN) — one of those three words, nothing else

Return a JSON array with exactly one object per variable listed above, in
the same order. Return valid JSON only — no markdown fences, no
explanation before or after.""" + (
        f"\n\nFollow CDISC SDTMIG v{sdtmig_version} conventions for domain "
        f"assignment, variable naming, and controlled terminology."
        if sdtmig_version else ""
    )


def _normalize_confidence(value):
    confidence = str(value or "Low").strip().title()
    return confidence if confidence in CONFIDENCE_LEVELS else "Low"


def _extract_json_array(text):
    """Pull just the JSON array out of a model response, regardless of
    what surrounds it — a leading ```json fence (assembler.clean() only
    strips ```sas/```r/bare ```, so a real Claude response wrapping JSON
    in ```json left the literal word "json" sitting before the array and
    broke json.loads entirely — caught live, not hypothetical), a trailing
    fence, or stray prose before/after. The prompt always asks for a JSON
    ARRAY (never a bare object), so slicing from the first "[" to the last
    "]" is reliable here in a way it wouldn't be for arbitrary code."""
    start = text.find("[")
    end = text.rfind("]")
    if start == -1 or end == -1 or end <= start:
        raise ValueError("no JSON array found in model response")
    return text[start:end + 1]


def map_dataset(dataset_name, var_meta, use_api=True, sdtmig_version=None):
    """One model call for this dataset's whole variable list. Returns one
    row per variable in var_meta, in order — never fewer, even if the
    model's response is missing one or fails outright (a variable the
    model didn't map cleanly becomes a Low-confidence UNKNOWN row for a
    human to look at, not a silently dropped one)."""
    prompt = build_mapping_prompt(dataset_name, var_meta, sdtmig_version=sdtmig_version)
    try:
        raw = generate_api(prompt) if use_api else generate_local(prompt)
        mappings = json.loads(_extract_json_array(raw))
        by_var = {m.get("source_var"): m for m in mappings if isinstance(m, dict)}
    except Exception as e:
        print(f"    Warning: mapping call failed for {dataset_name}: {e}")
        by_var = {}

    rows = []
    for v in var_meta:
        m = by_var.get(v["variable"], {})
        rows.append({
            "source_dataset": dataset_name,
            "source_variable": v["variable"],
            "source_label": v["label"],
            "sample_values": ", ".join(v["samples"]),
            "sdtm_domain": (m.get("sdtm_domain") or "UNKNOWN").strip().upper(),
            "sdtm_var": (m.get("sdtm_var") or "").strip().upper(),
            "qualifier": (m.get("qualifier") or "").strip(),
            "mapping_logic": m.get("mapping_logic") or
                             ("Model did not return a mapping for this variable."
                              if v["variable"] not in by_var else ""),
            "confidence": _normalize_confidence(m.get("confidence")),
        })
    return rows


# ── Excel export ─────────────────────────────────────────────────────

def write_automap_xlsx(rows, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="Arial", size=10)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    # Excel's own standard "Good/Neutral/Bad" cell-style colors — deliberately
    # distinct from the pale blue/green used for domain provenance elsewhere
    # in this app's other review workbooks, so confidence and domain-type
    # tinting are never visually confused.
    confidence_fill = {
        "High": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Low": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    # ── Sheet 1: Review ─────────────────────────────────────────────

    ws = wb.active
    ws.title = "Review"

    review_headers = ["Source Dataset", "Source Variable", "Source Label",
                      "Sample Values", "SDTM Domain", "SDTM Variable",
                      "Mapping Logic", "Confidence", "Review Status"]
    review_widths = [16, 16, 28, 40, 12, 14, 45, 12, 14]

    for ci, (h, w) in enumerate(zip(review_headers, review_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, header_align, thin_border
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, r in enumerate(rows, 2):
        vals = [r["source_dataset"], r["source_variable"], r["source_label"],
                r["sample_values"], r["sdtm_domain"], r["sdtm_var"],
                r["mapping_logic"], r["confidence"], ""]
        fill = confidence_fill[r["confidence"]]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font, cell.fill, cell.border = body_font, fill, thin_border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(review_headers))}{len(rows) + 1}"

    # ── Sheet 2: By Domain (sdtm_spec_builder.py-compatible) ─────────

    ws2 = wb.create_sheet(title="By Domain")

    domain_headers = ["Domain", "Variable", "CRF Label", "Codelist",
                      "Qualifier", "CRF Page", "Page Title", "SUPP?", "Review Status"]
    domain_widths = [12, 18, 30, 20, 20, 16, 40, 8, 15]

    for ci, (h, w) in enumerate(zip(domain_headers, domain_widths), 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, header_align, thin_border
        ws2.column_dimensions[get_column_letter(ci)].width = w

    mapped_rows = [r for r in rows if r["sdtm_domain"] != "UNKNOWN" and r["sdtm_var"]]
    for ri, r in enumerate(mapped_rows, 2):
        vals = [
            r["sdtm_domain"],
            r["sdtm_var"],
            r["source_label"] or r["source_variable"],
            "",
            r["qualifier"],
            r["source_dataset"],
            r["mapping_logic"],
            "Yes" if r["sdtm_domain"].startswith("SUPP") else "",
            "",
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.font, cell.border = body_font, thin_border

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(domain_headers))}{len(mapped_rows) + 1}"

    wb.save(output_path)

    n_high = sum(1 for r in rows if r["confidence"] == "High")
    n_med = sum(1 for r in rows if r["confidence"] == "Medium")
    n_low = sum(1 for r in rows if r["confidence"] == "Low")
    n_unknown = sum(1 for r in rows if r["sdtm_domain"] == "UNKNOWN")
    print(f"\nWrote {output_path}")
    print(f"  Review sheet   : {len(rows)} variables ({n_high} High, {n_med} Medium, {n_low} Low confidence)")
    print(f"  By Domain sheet: {len(mapped_rows)} mapped variables ({n_unknown} unmapped/UNKNOWN excluded)")
    print(f"  Review \"By Domain\", delete rejected rows, then:")
    print(f"    python sdtm_spec_builder.py {output_path} --output sdtm_spec_draft.xlsx")


# ── Main pipeline ────────────────────────────────────────────────────

def run_automapper(raw_data_dir, output_path, use_api=True, sdtmig_version=None, cancel_event=None):
    """cancel_event: optional threading.Event, checked once per dataset (a
    mapping call runs one whole dataset at a time, so between datasets is
    the natural granularity — same convention as assemble_adsl's
    cancel_event, which checks once per variable) so a web UI's Abort can
    stop this cleanly between datasets instead of only after all of them."""
    print(f"Reading raw datasets from: {raw_data_dir}")
    files = sorted(
        f for f in os.listdir(raw_data_dir)
        if f.lower().endswith((".sas7bdat", ".csv"))
    )
    if not files:
        raise ValueError(f"No .sas7bdat or .csv files found in {raw_data_dir}")
    print(f"  {len(files)} dataset(s): {', '.join(files)}")

    all_rows = []
    for fname in files:
        if cancel_event is not None and cancel_event.is_set():
            print(f"\n  Aborted before {fname}")
            break

        dataset_name = os.path.splitext(fname)[0]
        path = os.path.join(raw_data_dir, fname)
        print(f"\n  Dataset: {dataset_name} ({fname})")

        df, labels, formats = read_dataset(path)
        var_meta = extract_variable_metadata(df, labels, formats)
        print(f"    {len(var_meta)} variables, {len(df)} rows")

        mapped = map_dataset(dataset_name, var_meta, use_api=use_api, sdtmig_version=sdtmig_version)
        n_high = sum(1 for m in mapped if m["confidence"] == "High")
        n_med = sum(1 for m in mapped if m["confidence"] == "Medium")
        n_low = sum(1 for m in mapped if m["confidence"] == "Low")
        print(f"    Mapped: {n_high} High, {n_med} Medium, {n_low} Low confidence")
        all_rows.extend(mapped)

    write_automap_xlsx(all_rows, output_path)
    return all_rows


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Propose SDTM mappings directly from raw clinical datasets (no spec needed).")
    parser.add_argument("raw_data_dir", help="Folder of raw .sas7bdat or .csv files")
    parser.add_argument("--output", "-o", default="sdtm_automap.xlsx")
    parser.add_argument("--offline", action="store_true",
                        help="Use local Ollama instead of the Claude API")
    parser.add_argument("--sdtmig-version", default=None,
                        help="CDISC SDTMIG version (e.g. 3.4) to target — appended to "
                             "the mapping prompt so domain/variable choices follow that "
                             "version's conventions (default: no version instruction)")
    args = parser.parse_args()
    run_automapper(args.raw_data_dir, args.output, use_api=not args.offline,
                   sdtmig_version=args.sdtmig_version)
