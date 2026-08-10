"""
build_sample_adam_bds_spec.py - generate a sample adam_bds_spec.xlsx for
ADVS and ADAE, for adam_bds_assembler.py to read. Mirrors the existing
build_sample_acrf.py / build_sample_protocol.py pattern: generate once,
commit the output as a fixture.

Rows are in dependency order within each dataset (PARAMCD before AVAL
before ANL01FL, etc.) — adam_bds_assembler.py builds each dataset's
variables in spec row order, not alphabetically, since a later variable
often needs an earlier one already derived in the same data step.

Derivation rules only reference columns real sdtm_spec_draft.xlsx VS/AE
sheets actually have (verified via read_sdtm_domain_vars against this
repo's own sdtm_spec_draft.xlsx) — ADVS: VSTESTCD, VSSTRESN; ADAE:
AESTDTC, AEENDTC, AESEV, AEREL.
"""

import argparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ADVS_ROWS = [
    ("ADVS", "PARAMCD", "Parameter Code", "Char", 8, "VSTESTCD",
     "Set PARAMCD to VSTESTCD (already a short code, e.g. SYSBP, DIABP, PULSE)."),
    ("ADVS", "PARAM", "Parameter", "Char", 40, "VSTESTCD",
     "Set PARAM using this lookup on VSTESTCD: SYSBP='Systolic Blood Pressure "
     "(mmHg)'; DIABP='Diastolic Blood Pressure (mmHg)'; PULSE='Pulse Rate "
     "(beats/min)'; otherwise missing."),
    ("ADVS", "AVAL", "Analysis Value", "Num", 8, "VSSTRESN",
     "Set AVAL to VSSTRESN, the standardized numeric result."),
    ("ADVS", "ANL01FL", "Analysis Flag 01", "Char", 1, "Derived",
     "Set ANL01FL to 'Y' when AVAL is not missing; otherwise leave null."),
]

ADAE_ROWS = [
    ("ADAE", "ASTDT", "Analysis Start Date", "Num", 8, "AESTDTC",
     "Convert AESTDTC (an ISO 8601 date/time string) to a numeric SAS date "
     "using the E8601DA informat on the first 10 characters."),
    ("ADAE", "AENDT", "Analysis End Date", "Num", 8, "AEENDTC",
     "Convert AEENDTC (an ISO 8601 date/time string) to a numeric SAS date "
     "using the E8601DA informat on the first 10 characters."),
    ("ADAE", "AESEVN", "Severity/Intensity (N)", "Num", 8, "AESEV",
     "Set AESEVN to the numeric severity code from AESEV: MILD=1, "
     "MODERATE=2, SEVERE=3; otherwise leave missing."),
    ("ADAE", "AERELFL", "Relationship to Treatment Flag", "Char", 1, "AEREL",
     "Set AERELFL to 'Y' if AEREL indicates any relationship to study "
     "treatment (i.e. AEREL is not missing and not 'NOT RELATED'); "
     "otherwise 'N'."),
]


def build_adam_bds_spec(path="adam_bds_spec.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Variables"

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="Arial", size=10)
    advs_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    adae_fill = PatternFill(start_color="FDF3E3", end_color="FDF3E3", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["Dataset", "Variable", "Label", "Type", "Length", "Source", "Derivation"]
    widths = [10, 12, 28, 8, 8, 12, 55]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, header_align, thin_border
        ws.column_dimensions[get_column_letter(ci)].width = w

    all_rows = ADVS_ROWS + ADAE_ROWS
    for ri, row in enumerate(all_rows, 2):
        fill = advs_fill if row[0] == "ADVS" else adae_fill
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font, cell.fill, cell.border = body_font, fill, thin_border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(all_rows) + 1}"

    wb.save(path)
    print(f"Wrote {path}")
    print(f"  ADVS: {len(ADVS_ROWS)} variables, ADAE: {len(ADAE_ROWS)} variables")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate a sample ADaM BDS spec (ADVS + ADAE).")
    parser.add_argument("--output", "-o", default="adam_bds_spec.xlsx")
    args = parser.parse_args()
    build_adam_bds_spec(args.output)
