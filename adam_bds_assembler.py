"""
adam_bds_assembler.py - Generate ADaM BDS datasets (ADVS, ADLB, ADAE, ADCM,
ADEFF) from a hand-authored spec, an SDTM spec, and mock TLF shells.

Different path from bds_assembler.py: that one is deterministic (PARAMCD/
PARAM/BASE/CHG/ANL01FL built as fixed SAS/R templates, PARAMCD list read
straight off acrf_metadata.xlsx's Qualifier column, no model calls at all)
— it's the fast, no-spec-needed path for straightforward Findings/Events
domains. This module is for when a study needs REAL derivation logic per
variable (custom analysis flags, non-standard shift categories, a
condition a mock shell defines) and a human has written a proper spec for
it — the same "spec -> per-variable Writer/Improver/Reviewer" shape as
ADSL, just for BDS. Complementary, not a replacement; pick whichever the
study actually needs.

Three inputs:
  1. adam_bds_spec.xlsx  — Variables sheet: Dataset, Variable, Label, Type,
     Length, Source, Derivation. One row per variable per dataset — several
     datasets can share one spec file, distinguished by Dataset.
  2. --sdtm sdtm_spec_draft.xlsx — sdtm_spec_builder.py's own output (read,
     never written, by this module). Only used to tell the Writer which raw
     SDTM columns actually exist for the source domain (e.g. VSTESTCD,
     VSSTRESN, VISITNUM for VS) — real variables to reference, not invented
     ones.
  3. --shells mock_shells/ — a folder of Shell_Meta/Shell_Rows workbooks,
     the SAME schema tlf_assembler.py already reads (see read_shell) —
     existing sample_shell_demographics.xlsx / sample_shell_ae.xlsx work
     here unchanged. Scanned for whichever shell's Shell_Rows shape and
     source/numerator dataset match the ADaM dataset being built:
       - Findings-shape (paramcd/param columns, e.g. a VS shift shell) ->
         which PARAMCDs are needed, baseline visit, shift-category cutoffs
       - Events-shape (row_label/condition columns, e.g. the AE summary
         shell) -> which conditions (candidate analysis flags) a table
         actually needs
     A dataset with no matching shell still builds fine from the spec
     alone — shells are context, not a hard requirement.

Domain-class-aware prompts (see build_bds_prompt): ADVS/ADLB/ADEFF get the
Findings framing (PARAMCD/PARAM/AVAL/BASE/CHG/PCHG, ADLB also ANRLO/ANRHI);
ADAE/ADCM get the Events framing (one record per event, no baseline).

Same three-agent pipeline as everywhere else in this app: draft (Writer,
generate_api or generate_local under --offline) -> improve (a second call,
same routing) -> review (review_sas, generator.py's Reviewer-role alias).
Every block is BEGIN/END-wrapped like every other assembler in this repo,
so spec_differ.py/spec_patcher.py work on this output unchanged too.

CLI usage:
  python adam_bds_assembler.py adam_bds_spec.xlsx --sdtm sdtm_spec_draft.xlsx --shells mock_shells/ --output adam_programs/
  python adam_bds_assembler.py adam_bds_spec.xlsx --dataset ADVS --output adam_programs/
  python adam_bds_assembler.py adam_bds_spec.xlsx --output adam_programs/ --offline
"""

import argparse
import os

import pandas as pd
import openpyxl

from generator import generate_api, generate_local, review_sas
from runlog import log_run
from assembler import clean

BEGIN = "/*-- BEGIN {var} --*/"
END = "/*-- END {var} --*/"

FINDINGS_DATASETS = {"ADVS", "ADLB", "ADEFF"}
EVENTS_DATASETS = {"ADAE", "ADCM"}


def wrap(var, code):
    return f"{BEGIN.format(var=var)}\n{code}\n{END.format(var=var)}\n"


# ── Reading the three inputs ─────────────────────────────────────────

def read_bds_spec(spec_path):
    """adam_bds_spec.xlsx's Variables sheet -> a DataFrame with Dataset,
    Variable, Label, Type, Length, Source, Derivation."""
    return pd.read_excel(spec_path, sheet_name="Variables")


def read_sdtm_domain_vars(sdtm_spec_path, domain):
    """The variable list for one domain sheet of an sdtm_spec_draft.xlsx —
    what this module tells the Writer is actually available to reference
    from the SDTM source, so it doesn't invent columns. Returns [] (not an
    error) if the file or sheet doesn't exist — a spec-only build without
    an SDTM cross-check is still valid, just less guided.

    sdtm_spec_builder.write_sdtm_spec puts a "Domain: VS (Findings)" title
    on row 1 (row 2 for SUPP--) before the real header row — the header
    isn't always row 1, so this searches for the row whose first cell is
    literally "Variable" rather than assuming a fixed position."""
    if not sdtm_spec_path or not os.path.exists(sdtm_spec_path):
        return []
    try:
        wb = openpyxl.load_workbook(sdtm_spec_path, read_only=True, data_only=True)
        if domain not in wb.sheetnames:
            return []
        ws = wb[domain]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return []

    header_idx = next((i for i, r in enumerate(rows) if r and r[0] == "Variable"), None)
    if header_idx is None:
        return []
    return [r[0] for r in rows[header_idx + 1:] if r and r[0]]


def read_shell(path):
    """Same Shell_Meta/Shell_Rows schema tlf_assembler.py's _read_shell
    reads — existing sample shells work here with no changes."""
    meta_df = pd.read_excel(path, sheet_name="Shell_Meta")
    meta = dict(zip(meta_df["Field"], meta_df["Value"]))
    rows_df = pd.read_excel(path, sheet_name="Shell_Rows")
    return meta, rows_df


def classify_shell(rows_df):
    cols = set(rows_df.columns)
    if {"paramcd", "param"}.issubset(cols):
        return "findings"
    if {"row_label", "condition"}.issubset(cols):
        return "events"
    if {"adam_var", "stat_type"}.issubset(cols):
        return "demographics"  # ADSL-sourced — not relevant to BDS generation
    return "unknown"


def _extract_findings_requirements(meta, rows_df):
    paramcds = []
    for _, r in rows_df.iterrows():
        entry = {"paramcd": r["paramcd"], "param": r.get("param") or r["paramcd"]}
        if "low_cutoff" in rows_df.columns and not pd.isna(r.get("low_cutoff")):
            entry["low"] = r["low_cutoff"]
        if "high_cutoff" in rows_df.columns and not pd.isna(r.get("high_cutoff")):
            entry["high"] = r["high_cutoff"]
        paramcds.append(entry)
    return {
        "kind": "findings",
        "paramcds": paramcds,
        "baseline_visit": meta.get("baseline_visit") or "BASELINE",
        "population": meta.get("population", ""),
        "column_var": meta.get("column_var", ""),
    }


def _extract_events_requirements(meta, rows_df):
    conditions = []
    for _, r in rows_df.iterrows():
        cond = str(r.get("condition") or "").strip()
        if not cond or cond.lower() == "nan":
            continue
        conditions.append({"label": r["row_label"], "condition": cond})
    return {
        "kind": "events",
        "conditions": conditions,
        "teae_flag": meta.get("teae_flag", "TRTEMFL"),
        "population": meta.get("population", ""),
        "column_var": meta.get("column_var", ""),
    }


def gather_shell_requirements(shells_dir, dataset_code):
    """Scan every .xlsx in shells_dir; return the requirements dict from
    the first shell whose shape and source/numerator dataset match
    dataset_code, or {} if none do — a dataset with no matching shell
    still builds fine from the spec alone."""
    if not shells_dir or not os.path.isdir(shells_dir):
        return {}
    for fname in sorted(os.listdir(shells_dir)):
        if not fname.lower().endswith(".xlsx"):
            continue
        path = os.path.join(shells_dir, fname)
        try:
            meta, rows_df = read_shell(path)
        except Exception:
            continue
        kind = classify_shell(rows_df)
        if kind == "findings" and str(meta.get("source_dataset", "")).upper() == dataset_code.upper():
            return _extract_findings_requirements(meta, rows_df)
        if kind == "events" and str(meta.get("numer_dataset", "")).upper() == dataset_code.upper():
            return _extract_events_requirements(meta, rows_df)
    return {}


# ── Domain-class-aware prompts ───────────────────────────────────────

def _findings_context(dataset_code, sdtm_vars, shell_reqs):
    lines = [
        f"{dataset_code} is a Findings-class BDS dataset: one row per subject per "
        "PARAMCD per visit, built directly from the raw SDTM findings domain "
        "already SET into this step (see Variable list below) — PARAMCD, PARAM, "
        "AVAL, BASE, CHG, PCHG, and every other derived variable are NOT pre-built; "
        "each is derived here by its own Derivation rule, in the order given, so a "
        "later variable (e.g. CHG) can reference an earlier one (e.g. BASE) only "
        "once this step has already derived it.",
    ]
    if dataset_code == "ADLB":
        lines.append("ADLB-specific: ANRLO/ANRHI (normal range bounds) are also "
                     "available for out-of-range and shift derivations.")
    if sdtm_vars:
        lines.append(f"Raw SDTM source columns available: {', '.join(sdtm_vars)}.")
    if shell_reqs.get("paramcds"):
        pnames = ", ".join(p["paramcd"] for p in shell_reqs["paramcds"])
        lines.append(f"The mock shell for this table needs these PARAMCDs: {pnames}.")
        ranged = [p for p in shell_reqs["paramcds"] if "low" in p or "high" in p]
        for p in ranged:
            lines.append(f"  {p['paramcd']}: normal range {p.get('low', '?')}-{p.get('high', '?')} "
                         "(for a shift/out-of-range derivation, if this variable is one).")
    if shell_reqs.get("baseline_visit"):
        lines.append(f"Baseline visit for BASE/CHG/PCHG: {shell_reqs['baseline_visit']}.")
    return "\n".join(lines)


def _events_context(dataset_code, sdtm_vars, shell_reqs):
    lines = [
        f"{dataset_code} is an Events-class BDS dataset: one row per event/record "
        "per subject — no baseline, no PARAMCD reshape. Only USUBJID and the raw "
        "SDTM source columns below are pre-built; ADSL treatment dates (TRTSDT/"
        "TRTEDT), numeric analysis dates (ASTDT/AENDT), and every other derived "
        "variable must each be derived here by its own Derivation rule, in the "
        "order given — if a variable's rule needs TRTSDT or ASTDT, an earlier "
        "spec row must derive it first (e.g. from an ADSL merge or a --DTC "
        "conversion) before a later row can reference it.",
    ]
    if dataset_code == "ADAE":
        lines.append("ADAE-specific: treatment-emergent flag is TRTEMFL; severity "
                     "grades come from AESEV (character) / AESEVN (numeric, "
                     "MILD=1/MODERATE=2/SEVERE=3).")
    if dataset_code == "ADCM":
        lines.append("ADCM-specific: ATC classification is CMCLAS; PREFL (prior), "
                     "CONFL (concomitant), and ONTRTFL (on-treatment) flag when the "
                     "medication was taken relative to treatment dates.")
    if sdtm_vars:
        lines.append(f"Raw SDTM source columns available: {', '.join(sdtm_vars)}.")
    if shell_reqs.get("conditions"):
        lines.append("The mock shell for this table needs these condition-based rows "
                     "(candidate analysis-flag logic):")
        for c in shell_reqs["conditions"]:
            lines.append(f'  "{c["label"]}": {c["condition"]}')
    return "\n".join(lines)


def build_bds_prompt(dataset_code, row, sdtm_vars, shell_reqs, known_vars):
    """One Writer prompt for one variable. Domain-class framing from
    _findings_context/_events_context; everything else (metadata, output
    rules) is shared."""
    is_findings = dataset_code in FINDINGS_DATASETS
    context = (_findings_context if is_findings else _events_context)(
        dataset_code, sdtm_vars, shell_reqs)

    length = row.get("Length", "")
    source = str(row.get("Source", "") or "").strip()
    source_line = f"Source: {source}\n" if source and source.lower() != "nan" else ""

    return f"""You are a senior clinical SAS programmer.
Write SAS 9.4 code to derive one variable in the {dataset_code} dataset.

{context}

Variable: {row['Variable']}
Label: {row['Label']}
Type: {row['Type']}, Length: {length}
{source_line}Derivation rule: {row['Derivation']}

Variables already available in this step: {', '.join(known_vars)}

Rules:
- Output ONLY the derivation logic statements (length, label, format, if/then, assignments).
- Do NOT include data, set, merge, or run statements - the code will be inserted into an existing data step.
- Do NOT repeat the variable metadata as comments. Output ONE brief comment line, then the code.
- Do NOT add any explanation before or after the code.
- Output plain SAS code only, no markdown fences.
- Do NOT invent format names, variable values, or codes not stated in the derivation rule.
- Reference variables by name only, never with a dataset prefix.
- CRITICAL: In SAS, missing numeric values are less than every number. Any numeric range chain MUST guard `missing(X)` first.
"""


def build_improve_prompt(code, dataset_code, row):
    return f"""You are a principal clinical SAS programmer with 15+ years of experience.
Rewrite the draft below for {dataset_code} variable {row['Variable']}
({row['Label']}) to sign-off quality.

Derivation rule: {row['Derivation']}

DRAFT CODE
{code}

Rewrite the code so that it:
- Correctly implements the derivation rule and nothing more
- Declares length and label; adds a format ONLY if the spec gives one
- Never invents values, codes, or dataset/variable names
- Is code a principal programmer would sign off on

Output ONLY the corrected SAS code with one brief comment line.
No explanation, no markdown fences.
"""


def build_review_prompt(code, known_vars):
    return f"""You are a senior clinical SAS programmer performing QC.

Review this generated SAS code block:

{code}

Variables available in this data step:
{", ".join(known_vars)}

Check ONLY for these issues:
1. References to variables NOT in the available list (hallucinated variables)
2. Statements that would fail in a data step (data/set/merge/run statements)
3. Character values assigned to numeric variables or vice versa

Reply with exactly one line:
PASS
or
FAIL: <short reason>

No other text.
"""


# ── Per-variable three-agent pipeline ────────────────────────────────

def build_variable_block(dataset_code, row, sdtm_vars, shell_reqs, known_vars, use_api=True):
    """Draft -> improve -> review for one variable. Returns the wrapped
    BEGIN/END block; a Reviewer FAIL is prepended as a QC FLAG comment,
    same convention as every other assembler in this app."""
    var = row["Variable"]
    print(f"  Writing: {var}")
    draft_fn = generate_api if use_api else generate_local
    draft = clean(draft_fn(build_bds_prompt(dataset_code, row, sdtm_vars, shell_reqs, known_vars)))

    print(f"  Improving: {var}")
    improved = clean(draft_fn(build_improve_prompt(draft, dataset_code, row)))

    print(f"  Reviewing: {var}")
    verdict = review_sas(build_review_prompt(improved, known_vars), mode="api" if use_api else "local").strip()

    block = improved
    if verdict.startswith("FAIL"):
        block = f"/* QC FLAG: {verdict} */\n{block}"

    return wrap(var, block)


# ── One dataset ───────────────────────────────────────────────────────

def build_dataset(dataset_code, spec, sdtm_spec_path, shells_dir, use_api=True):
    """Build one BDS dataset's SAS program from its spec rows."""
    rows = spec[spec["Dataset"].astype(str).str.upper() == dataset_code.upper()]
    if rows.empty:
        raise ValueError(f"No rows for Dataset={dataset_code!r} in the spec")

    sdtm_domain = dataset_code[2:]  # ADVS -> VS, ADAE -> AE, ADLB -> LB, ADCM -> CM
    sdtm_vars = read_sdtm_domain_vars(sdtm_spec_path, sdtm_domain)
    shell_reqs = gather_shell_requirements(shells_dir, dataset_code)
    if shell_reqs:
        print(f"  Shell requirements found ({shell_reqs['kind']})")

    ds = dataset_code.lower()
    header = [
        "/* ********************************* */",
        f"/* Program: {ds}_spec.sas */",
        "/* Generated by SpecGen (adam_bds_assembler.py) */",
        "/* ********************************* */",
        "",
        f"data {ds};",
        f"    set {sdtm_domain.lower()};  /* raw SDTM {sdtm_domain} domain — every ADaM-only",
        "       variable below (PARAMCD/AVAL/BASE/..., ADSL merges, date conversions,",
        "       analysis flags) is derived by its own spec row, not pre-built */",
        "",
    ]
    program = "\n".join(header) + "\n"

    # What this data step actually has, before any spec row runs: USUBJID plus
    # whatever raw columns the SDTM spec says this domain has. Each spec row's
    # own Variable name is appended as it's built, so row N can reference
    # anything rows 1..N-1 already derived (same discipline as ADSL's spec order).
    known_vars = ["USUBJID"] + [v for v in sdtm_vars if v != "USUBJID"]

    # Row order, not alphabetical: later variables often depend on earlier
    # ones (BASE needs PARAMCD/AVAL already set, CHG needs BASE, ...), so
    # the spec's own row order IS the dependency order — same convention
    # ADSL's assembler.py uses (main_step.sort_values("Order")), just
    # keyed on file order here since this spec has no Order column.
    for _, row in rows.iterrows():
        block = build_variable_block(dataset_code, row, sdtm_vars, shell_reqs, known_vars, use_api=use_api)
        program += block + "\n"
        known_vars.append(row["Variable"])

    program += "run;\n"
    return program, len(rows)


# ── Main pipeline ────────────────────────────────────────────────────

def run_adam_bds(spec_path, sdtm_spec_path=None, shells_dir=None, output_dir="adam_programs",
                 dataset=None, use_api=True):
    print(f"Reading ADaM BDS spec: {spec_path}")
    spec = read_bds_spec(spec_path)
    datasets = [dataset.upper()] if dataset else sorted(spec["Dataset"].astype(str).str.upper().unique())
    print(f"  Dataset(s): {', '.join(datasets)}")

    os.makedirs(output_dir, exist_ok=True)
    for dataset_code in datasets:
        print(f"\nBuilding {dataset_code}")
        program, n_vars = build_dataset(dataset_code, spec, sdtm_spec_path, shells_dir, use_api=use_api)

        # "_spec" suffix, not "{dataset}.sas" — bds_assembler.py already
        # writes adam_programs/advs.sas, adae.sas, etc. deterministically;
        # this is a genuinely different, spec-driven program for the same
        # dataset, not a replacement, and letting them share a filename
        # would mean whichever tool runs second silently clobbers the
        # other's (differently generated, differently reviewed) output.
        path = os.path.join(output_dir, f"{dataset_code.lower()}_spec.sas")
        with open(path, "w", encoding="utf-8") as f:
            f.write(program)
        print(f"  Wrote {path} ({n_vars} variables)")

        log_run(
            spec_path,
            f"adam_bds/{'api' if use_api else 'local'}",
            "claude-sonnet-4-5" if use_api else "qwen2.5-coder:7b",
            "claude-sonnet-4-5" if use_api else "qwen2.5-coder:7b",
            "claude-sonnet-4-5" if use_api else "qwen2.5-coder:7b",
            n_vars,
            path,
        )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate ADaM BDS datasets (ADVS/ADLB/ADAE/ADCM/ADEFF) from a spec, "
                    "SDTM spec, and mock shells.")
    parser.add_argument("spec", help="Path to adam_bds_spec.xlsx")
    parser.add_argument("--sdtm", default=None, help="Path to sdtm_spec_draft.xlsx")
    parser.add_argument("--shells", default=None, help="Folder of mock shell .xlsx files")
    parser.add_argument("--output", "-o", default="adam_programs")
    parser.add_argument("--dataset", "-d", default=None,
                        help="Build a single dataset (e.g. ADVS) instead of every dataset in the spec")
    parser.add_argument("--offline", action="store_true",
                        help="Use local Ollama instead of the Claude API")
    args = parser.parse_args()

    run_adam_bds(args.spec, sdtm_spec_path=args.sdtm, shells_dir=args.shells,
                output_dir=args.output, dataset=args.dataset, use_api=not args.offline)
