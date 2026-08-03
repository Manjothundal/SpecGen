"""
sdtm_assembler.py - Generate SDTM SAS programs from a draft SDTM spec.

Phase 5c in the SpecGen pipeline:
  sdtm_spec_draft.xlsx --> dm.sas, ae.sas, vs.sas, cm.sas, ...

Unlike the ADaM assembler (variable-by-variable), SDTM generates one
complete program per domain because the variables are interdependent
(e.g. VSTESTCD and VSORRES must be built together in the same data step).

Domain class determines the code structure:
  DM         : one row per subject, merge demographics from multiple sources
  Events     : one row per event (AE, DS, MH, DV) — read source, derive timing
  Interventions : one row per intervention (CM, EX) — read source, derive timing
  Findings   : one row per test per timepoint (VS, EG) — transpose or stack tests
  Findings About Events : one row per assessment (TU, TR, RS) — similar to Findings
  SUPP--     : vertical QNAM/QVAL structure from parent domain; appended into
               the parent domain's own .sas file (e.g. SUPPAE lives inside
               ae.sas), not written as a separate program

Uses the existing three-agent pipeline:
  Writer (Ollama or API) --> Improver (API) --> Reviewer (API)

Usage:
  # Generate all domains
  python sdtm_assembler.py sdtm_spec_draft.xlsx --output sdtm_programs/

  # Generate one domain
  python sdtm_assembler.py sdtm_spec_draft.xlsx --domain DM --output sdtm_programs/

  # Offline mode (Ollama only, no API)
  python sdtm_assembler.py sdtm_spec_draft.xlsx --offline --output sdtm_programs/

  # As a module
  from sdtm_assembler import generate_all_domains
  generate_all_domains("sdtm_spec_draft.xlsx", "sdtm_programs/")
"""

import os
import re
import sys
import argparse
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl

# Claude's responses (review verdicts, generated code) commonly contain
# Unicode punctuation (em-dashes, curly quotes) that Windows' default console
# codepage (cp1252) can't encode. Printing that text would otherwise crash
# this entire subprocess with UnicodeEncodeError the moment any domain's
# review happens to contain such a character — silently killing every
# domain's generation, not just the one that triggered it.
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

from config import WRITER, REVIEWER, LOCAL_MODEL, API_MODEL
from generator import generate_local, generate_api, review_sas
from improver import improve_block
from runlog import log_run
from macro_lookup import load_catalog, find_sdtm_macros

_macro_catalog = load_catalog()

# Domains generate concurrently (see generate_all_domains); log_run() appends
# a row to a shared CSV, and two threads finishing at the same instant could
# interleave their writes into one corrupted line. One lock, held only for
# the log write itself, keeps rows intact without limiting the actual
# API-bound work to one-at-a-time.
_LOG_LOCK = threading.Lock()


# ── Domain classification ───────────────────────────────────────────

EVENTS_DOMAINS = {"AE", "DS", "MH", "DV", "CE"}
INTERVENTIONS_DOMAINS = {"CM", "EC", "EX", "PR", "SU"}
FINDINGS_DOMAINS = {"VS", "LB", "EG", "PE", "QS", "SC", "DA", "MB", "MS", "PC", "PP"}
FINDINGS_ABOUT_EVENTS_DOMAINS = {"TU", "TR", "RS"}


def get_domain_class(domain):
    if domain.startswith("SUPP"):
        return "SUPP"
    if domain == "DM":
        return "DM"
    if domain in EVENTS_DOMAINS:
        return "Events"
    if domain in INTERVENTIONS_DOMAINS:
        return "Interventions"
    if domain in FINDINGS_DOMAINS:
        return "Findings"
    if domain in FINDINGS_ABOUT_EVENTS_DOMAINS:
        return "Findings About Events"
    return "General"


# ── Read SDTM spec ──────────────────────────────────────────────────

def read_domain_spec(xlsx_path, domain):
    """
    Read one domain sheet from sdtm_spec_draft.xlsx.
    Returns a list of variable dicts.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if domain not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[domain]

    # Find header row (skip domain label rows)
    header_row = None
    for row_num, row in enumerate(ws.iter_rows(values_only=False), 1):
        vals = [cell.value for cell in row]
        if "Variable" in vals and "Label" in vals:
            header_row = row_num
            break

    if header_row is None:
        wb.close()
        return []

    headers = [cell.value for cell in ws[header_row]]
    variables = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        record = dict(zip(headers, row))
        if record.get("Variable"):
            variables.append(record)

    wb.close()
    return variables


def list_domains(xlsx_path):
    """List all domain sheets in the spec (excluding Cover)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    domains = [s for s in wb.sheetnames if s != "Cover"]
    wb.close()
    return domains


# ── Prompt building per domain class ────────────────────────────────

def _var_table(variables):
    """Format variable list as a readable table for the prompt. Every one
    of the 6 domain-class prompt builders (x2 languages) calls this, so
    it's the single point where a per-variable Comment (free-text mapping
    notes/overrides, same spirit as the ADaM spec's Comment column added in
    prompt_builder.py) reaches the Writer — appended as its own list below
    the table rather than crammed into a fixed-width column, since comments
    can run much longer than Label/Origin/Codelist."""
    lines = []
    lines.append(f"{'Variable':<16} {'Label':<45} {'Type':<6} {'Len':<5} {'Origin':<10} {'Codelist'}")
    lines.append(f"{'-'*16} {'-'*45} {'-'*6} {'-'*5} {'-'*10} {'-'*20}")
    comment_lines = []
    for v in variables:
        var = str(v.get("Variable", ""))
        label = str(v.get("Label", ""))[:45]
        vtype = str(v.get("Type", ""))
        length = str(v.get("Length", ""))
        origin = str(v.get("Origin", ""))
        codelist = str(v.get("Codelist", "") or "")
        lines.append(f"{var:<16} {label:<45} {vtype:<6} {length:<5} {origin:<10} {codelist}")
        comment = str(v.get("Comment", "") or "").strip()
        if comment and comment.lower() != "nan":
            comment_lines.append(f"- {var}: {comment}")
    if comment_lines:
        lines.append("\nAdditional instructions (per variable):")
        lines.extend(comment_lines)
    return "\n".join(lines)


def build_dm_prompt(domain, variables, language="sas"):
    """Prompt for DM domain - one row per subject."""
    if language == "r":
        return _build_dm_prompt_r(domain, variables)
    var_table = _var_table(variables)
    return f"""You are a senior CDISC SDTM programmer. Generate a complete, production-quality SAS program
for the {domain} (Demographics) domain.

SDTM specification for {domain}:
{var_table}

Requirements:
- Read source data from raw. library (raw.dm, raw.ex as needed)
- Set STUDYID, DOMAIN as constants
- Derive USUBJID = catx('-', STUDYID, SITEID, SUBJID)
- Map CRF fields to SDTM variables per the spec
- Derive RFSTDTC (first dose date) and RFENDTC (last dose date) from EX domain
- Derive RFXSTDTC, RFXENDTC from exposure data
- ARM and ARMCD from randomization/CRF; ACTARM/ACTARMCD derived or same as planned
- Compute AGE from BRTHDTC and RFSTDTC if not directly collected
- COUNTRY from site-level metadata
- Apply proper lengths, labels, and formats per the spec
- Sort by STUDYID USUBJID
- Output to sdtm.{domain} with a KEEP statement listing all spec variables in order
- Include variable labels in a LABEL statement
- Add clear comments for each derivation section
- Use /*-- BEGIN {domain} --*/ and /*-- END {domain} --*/ markers around the main code block

Generate the complete SAS program. No explanations, just the code."""


def _build_dm_prompt_r(domain, variables):
    var_table = _var_table(variables)
    return f"""You are a senior CDISC SDTM programmer working in R (tidyverse). Generate a complete,
production-quality R script for the {domain} (Demographics) domain.

SDTM specification for {domain}:
{var_table}

Requirements:
- library(dplyr) at the top; read source data frames raw_dm, raw_ex (already in the R session,
  do NOT read/import them from a file)
- Set STUDYID, DOMAIN as constant columns
- Derive USUBJID = paste(STUDYID, SITEID, SUBJID, sep = "-")
- Map CRF fields to SDTM variables per the spec
- Derive RFSTDTC (first dose date) and RFENDTC (last dose date) from raw_ex
- Derive RFXSTDTC, RFXENDTC from exposure data
- ARM and ARMCD from randomization/CRF; ACTARM/ACTARMCD derived or same as planned
- Compute AGE from BRTHDTC and RFSTDTC if not directly collected
- COUNTRY from site-level metadata
- Assign the final result to a data frame named dm
- Sort with arrange(STUDYID, USUBJID)
- select() only the spec variables, in order, at the end — this is R's equivalent of a KEEP statement
- R has no length/label/format statements — do NOT emit them; use a comment instead to note each
  variable's label from the spec
- For ISO 8601 date strings (--DTC variables), parse with as.Date(substr(dtc, 1, 10)) — do not
  invent other date formats
- Add clear comments for each derivation section
- Use # -- BEGIN {domain} -- # and # -- END {domain} -- # markers around the main code block

Generate the complete R script. No explanations, just the code, no markdown fences."""


def build_events_prompt(domain, variables, language="sas"):
    """Prompt for Events domains (AE, DS, MH, DV) - one row per event."""
    if language == "r":
        return _build_events_prompt_r(domain, variables)
    var_table = _var_table(variables)
    domain_lower = domain.lower()
    prefix = domain[:2]
    return f"""You are a senior CDISC SDTM programmer. Generate a complete, production-quality SAS program
for the {domain} domain (Events class - one row per event per subject).

SDTM specification for {domain}:
{var_table}

Requirements:
- Read source data from raw.{domain_lower} (and raw.dm for STUDYID/USUBJID if needed)
- Set DOMAIN = '{domain}'
- Derive {prefix}SEQ as a sequence number within each subject (by USUBJID)
- Derive USUBJID = catx('-', STUDYID, SITEID, SUBJID) or merge from DM
- Map --TERM, --DECOD from source verbatim and coded terms
- Map --STDTC, --ENDTC from source start/end dates (ISO 8601 character format)
- Derive --STDY, --ENDY as study day relative to RFSTDTC from DM
- Derive EPOCH based on date relative to treatment period
- Apply --BODSYS from MedDRA/WHO coding if applicable
- Apply --CAT, --SCAT from source categories
- Handle domain-specific variables (severity for AE, disposition terms for DS, etc.)
- Apply proper lengths, labels, and formats per the spec
- Sort by STUDYID USUBJID {prefix}SEQ
- Output to sdtm.{domain} with a KEEP statement listing all spec variables in order
- Include variable labels in a LABEL statement
- Add clear comments for each derivation section
- Use /*-- BEGIN {domain} --*/ and /*-- END {domain} --*/ markers

Generate the complete SAS program. No explanations, just the code."""


def _build_events_prompt_r(domain, variables):
    var_table = _var_table(variables)
    domain_lower = domain.lower()
    prefix = domain[:2]
    return f"""You are a senior CDISC SDTM programmer working in R (tidyverse). Generate a complete,
production-quality R script for the {domain} domain (Events class - one row per event per subject).

SDTM specification for {domain}:
{var_table}

Requirements:
- library(dplyr) at the top; read source data frames raw_{domain_lower} and dm (both already in the
  R session, do NOT read/import them from a file)
- Set DOMAIN = "{domain}"
- Derive {prefix}SEQ as a sequence number within each subject: group_by(USUBJID) |>
  mutate({prefix}SEQ = row_number()) — NA-safe, no proc-sort-then-first. equivalent needed
- Derive USUBJID = paste(STUDYID, SITEID, SUBJID, sep = "-") or join from dm
- Map --TERM, --DECOD from source verbatim and coded terms
- Map --STDTC, --ENDTC from source start/end dates (ISO 8601 character format)
- Derive --STDY, --ENDY as study day relative to RFSTDTC from dm
- Derive EPOCH based on date relative to treatment period
- Apply --BODSYS from MedDRA/WHO coding if applicable
- Apply --CAT, --SCAT from source categories
- Handle domain-specific variables (severity for AE, disposition terms for DS, etc.)
- Assign the final result to a data frame named {domain_lower}
- Sort with arrange(STUDYID, USUBJID, {prefix}SEQ)
- select() only the spec variables, in order, at the end — this is R's equivalent of a KEEP statement
- R has no length/label/format statements — do NOT emit them; use a comment instead to note each
  variable's label from the spec
- For ISO 8601 date strings (--DTC variables), parse with as.Date(substr(dtc, 1, 10)) — do not
  invent other date formats
- Add clear comments for each derivation section
- Use # -- BEGIN {domain} -- # and # -- END {domain} -- # markers

Generate the complete R script. No explanations, just the code, no markdown fences."""


def build_interventions_prompt(domain, variables, language="sas"):
    """Prompt for Interventions domains (CM, EX) - one row per intervention."""
    if language == "r":
        return _build_interventions_prompt_r(domain, variables)
    var_table = _var_table(variables)
    domain_lower = domain.lower()
    prefix = domain[:2]
    return f"""You are a senior CDISC SDTM programmer. Generate a complete, production-quality SAS program
for the {domain} domain (Interventions class - one row per intervention per subject).

SDTM specification for {domain}:
{var_table}

Requirements:
- Read source data from raw.{domain_lower} (and raw.dm for STUDYID/USUBJID if needed)
- Set DOMAIN = '{domain}'
- Derive {prefix}SEQ as a sequence number within each subject (by USUBJID)
- Derive USUBJID = catx('-', STUDYID, SITEID, SUBJID) or merge from DM
- Map --TRT (reported treatment name) and --DECOD (standardized name) from source
- Map --DOSE, --DOSU, --DOSFRQ, --ROUTE from source dosing fields
- Map --STDTC, --ENDTC from source start/end dates (ISO 8601 character format)
- Derive --STDY, --ENDY as study day relative to RFSTDTC from DM
- Derive EPOCH based on date relative to treatment period
- Apply --CAT from source categories
- Apply proper lengths, labels, and formats per the spec
- Sort by STUDYID USUBJID {prefix}SEQ
- Output to sdtm.{domain} with a KEEP statement listing all spec variables in order
- Include variable labels in a LABEL statement
- Add clear comments for each derivation section
- Use /*-- BEGIN {domain} --*/ and /*-- END {domain} --*/ markers

Generate the complete SAS program. No explanations, just the code."""


def _build_interventions_prompt_r(domain, variables):
    var_table = _var_table(variables)
    domain_lower = domain.lower()
    prefix = domain[:2]
    return f"""You are a senior CDISC SDTM programmer working in R (tidyverse). Generate a complete,
production-quality R script for the {domain} domain (Interventions class - one row per
intervention per subject).

SDTM specification for {domain}:
{var_table}

Requirements:
- library(dplyr) at the top; read source data frames raw_{domain_lower} and dm (both already in
  the R session, do NOT read/import them from a file)
- Set DOMAIN = "{domain}"
- Derive {prefix}SEQ as a sequence number within each subject: group_by(USUBJID) |>
  mutate({prefix}SEQ = row_number())
- Derive USUBJID = paste(STUDYID, SITEID, SUBJID, sep = "-") or join from dm
- Map --TRT (reported treatment name) and --DECOD (standardized name) from source
- Map --DOSE, --DOSU, --DOSFRQ, --ROUTE from source dosing fields
- Map --STDTC, --ENDTC from source start/end dates (ISO 8601 character format)
- Derive --STDY, --ENDY as study day relative to RFSTDTC from dm
- Derive EPOCH based on date relative to treatment period
- Apply --CAT from source categories
- Assign the final result to a data frame named {domain_lower}
- Sort with arrange(STUDYID, USUBJID, {prefix}SEQ)
- select() only the spec variables, in order, at the end — this is R's equivalent of a KEEP statement
- R has no length/label/format statements — do NOT emit them; use a comment instead to note each
  variable's label from the spec
- For ISO 8601 date strings (--DTC variables), parse with as.Date(substr(dtc, 1, 10)) — do not
  invent other date formats
- Add clear comments for each derivation section
- Use # -- BEGIN {domain} -- # and # -- END {domain} -- # markers

Generate the complete R script. No explanations, just the code, no markdown fences."""


def build_findings_prompt(domain, variables, language="sas"):
    """Prompt for Findings domains (VS, EG) - one row per test per timepoint."""
    if language == "r":
        return _build_findings_prompt_r(domain, variables)
    var_table = _var_table(variables)
    domain_lower = domain.lower()
    prefix = domain[:2]
    return f"""You are a senior CDISC SDTM programmer. Generate a complete, production-quality SAS program
for the {domain} domain (Findings class - one row per test per timepoint per subject).

SDTM specification for {domain}:
{var_table}

Requirements:
- Read source data from raw.{domain_lower} (and raw.dm for STUDYID/USUBJID if needed)
- Set DOMAIN = '{domain}'
- Derive {prefix}SEQ as a sequence number within each subject (by USUBJID)
- Derive USUBJID = catx('-', STUDYID, SITEID, SUBJID) or merge from DM
- Each test parameter becomes one row: set {prefix}TESTCD and {prefix}TEST per measurement
- If source is wide format (one column per test), transpose to vertical (one row per test)
- If source is already vertical, map directly
- Map {prefix}ORRES (original result as character), {prefix}ORRESU (original units) from source
- Derive {prefix}STRESC (standardized character result) and {prefix}STRESN (numeric result)
- Derive {prefix}STRESU (standard units) - apply unit conversions if needed
- Map VISITNUM, VISIT from source visit data
- Map {prefix}DTC from source collection date (ISO 8601 character format)
- Derive {prefix}DY as study day relative to RFSTDTC from DM
- Handle {prefix}STAT = 'NOT DONE' and {prefix}REASND for missing measurements
- Apply proper lengths, labels, and formats per the spec
- Sort by STUDYID USUBJID {prefix}TESTCD VISITNUM {prefix}DTC
- Output to sdtm.{domain} with a KEEP statement listing all spec variables in order
- Include variable labels in a LABEL statement
- Add clear comments for each derivation section
- Use /*-- BEGIN {domain} --*/ and /*-- END {domain} --*/ markers

Generate the complete SAS program. No explanations, just the code."""


def _build_findings_prompt_r(domain, variables):
    var_table = _var_table(variables)
    domain_lower = domain.lower()
    prefix = domain[:2]
    return f"""You are a senior CDISC SDTM programmer working in R (tidyverse). Generate a complete,
production-quality R script for the {domain} domain (Findings class - one row per test per
timepoint per subject).

SDTM specification for {domain}:
{var_table}

Requirements:
- library(dplyr), library(tidyr) at the top; read source data frame raw_{domain_lower} and dm
  (both already in the R session, do NOT read/import them from a file)
- Set DOMAIN = "{domain}"
- Derive {prefix}SEQ as a sequence number within each subject: group_by(USUBJID) |>
  mutate({prefix}SEQ = row_number())
- Derive USUBJID = paste(STUDYID, SITEID, SUBJID, sep = "-") or join from dm
- Each test parameter becomes one row: set {prefix}TESTCD and {prefix}TEST per measurement
- If source is wide format (one column per test), pivot_longer() to vertical (one row per test)
- If source is already vertical, map directly
- Map {prefix}ORRES (original result as character), {prefix}ORRESU (original units) from source
- Derive {prefix}STRESC (standardized character result) and {prefix}STRESN (numeric result)
- Derive {prefix}STRESU (standard units) - apply unit conversions if needed
- Map VISITNUM, VISIT from source visit data
- Map {prefix}DTC from source collection date (ISO 8601 character format)
- Derive {prefix}DY as study day relative to RFSTDTC from dm
- Handle {prefix}STAT = "NOT DONE" and {prefix}REASND for missing measurements
- Assign the final result to a data frame named {domain_lower}
- Sort with arrange(STUDYID, USUBJID, {prefix}TESTCD, VISITNUM, {prefix}DTC)
- select() only the spec variables, in order, at the end — this is R's equivalent of a KEEP statement
- R has no length/label/format statements — do NOT emit them; use a comment instead to note each
  variable's label from the spec
- For ISO 8601 date strings (--DTC variables), parse with as.Date(substr(dtc, 1, 10)) — do not
  invent other date formats
- Add clear comments for each derivation section
- Use # -- BEGIN {domain} -- # and # -- END {domain} -- # markers

Generate the complete R script. No explanations, just the code, no markdown fences."""


def build_fae_prompt(domain, variables, language="sas"):
    """Prompt for Findings About Events domains (TU, TR, RS)."""
    if language == "r":
        return _build_fae_prompt_r(domain, variables)
    var_table = _var_table(variables)
    domain_lower = domain.lower()
    prefix = domain[:2]
    return f"""You are a senior CDISC SDTM programmer. Generate a complete, production-quality SAS program
for the {domain} domain (Findings About Events class - tumor/response assessments).

SDTM specification for {domain}:
{var_table}

Requirements:
- Read source data from raw.{domain_lower} (and raw.dm for STUDYID/USUBJID if needed)
- Set DOMAIN = '{domain}'
- Derive {prefix}SEQ as a sequence number within each subject (by USUBJID)
- Derive USUBJID from DM
- Each assessment becomes one row: set {prefix}TESTCD and {prefix}TEST per measurement
- Map {prefix}ORRES (original result), {prefix}STRESC, {prefix}STRESN from source
- Map {prefix}EVAL (evaluator: INVESTIGATOR or INDEPENDENT ASSESSOR)
- Map {prefix}LNKID for linking between TU/TR/RS domains
- Map VISITNUM, VISIT, {prefix}DTC from source
- Derive {prefix}DY relative to RFSTDTC from DM
- Apply {prefix}CAT for assessment criteria (e.g. RECIST 1.1)
- Apply proper lengths, labels, and formats per the spec
- Sort by STUDYID USUBJID {prefix}TESTCD VISITNUM {prefix}DTC
- Output to sdtm.{domain} with a KEEP statement
- Include variable labels in a LABEL statement
- Use /*-- BEGIN {domain} --*/ and /*-- END {domain} --*/ markers

Generate the complete SAS program. No explanations, just the code."""


def _build_fae_prompt_r(domain, variables):
    var_table = _var_table(variables)
    domain_lower = domain.lower()
    prefix = domain[:2]
    return f"""You are a senior CDISC SDTM programmer working in R (tidyverse). Generate a complete,
production-quality R script for the {domain} domain (Findings About Events class - tumor/response
assessments).

SDTM specification for {domain}:
{var_table}

Requirements:
- library(dplyr) at the top; read source data frames raw_{domain_lower} and dm (both already in
  the R session, do NOT read/import them from a file)
- Set DOMAIN = "{domain}"
- Derive {prefix}SEQ as a sequence number within each subject: group_by(USUBJID) |>
  mutate({prefix}SEQ = row_number())
- Derive USUBJID from dm (join)
- Each assessment becomes one row: set {prefix}TESTCD and {prefix}TEST per measurement
- Map {prefix}ORRES (original result), {prefix}STRESC, {prefix}STRESN from source
- Map {prefix}EVAL (evaluator: INVESTIGATOR or INDEPENDENT ASSESSOR)
- Map {prefix}LNKID for linking between TU/TR/RS domains
- Map VISITNUM, VISIT, {prefix}DTC from source
- Derive {prefix}DY relative to RFSTDTC from dm
- Apply {prefix}CAT for assessment criteria (e.g. RECIST 1.1)
- Assign the final result to a data frame named {domain_lower}
- Sort with arrange(STUDYID, USUBJID, {prefix}TESTCD, VISITNUM, {prefix}DTC)
- select() only the spec variables, in order, at the end — this is R's equivalent of a KEEP statement
- R has no length/label/format statements — do NOT emit them; use a comment instead to note each
  variable's label from the spec
- Use # -- BEGIN {domain} -- # and # -- END {domain} -- # markers

Generate the complete R script. No explanations, just the code, no markdown fences."""


def build_supp_prompt(domain, variables, language="sas"):
    """Prompt for SUPP-- domains - vertical QNAM/QVAL structure."""
    if language == "r":
        return _build_supp_prompt_r(domain, variables)
    var_table = _var_table(variables)
    parent = domain.replace("SUPP", "")
    parent_lower = parent.lower()

    # Separate structural vars from qualifier values
    structural = [v for v in variables if v.get("Variable") in
                  ("STUDYID", "RDOMAIN", "USUBJID", "IDVAR", "IDVARVAL",
                   "QNAM", "QLABEL", "QVAL", "QORIG", "QEVAL")]
    qualifiers = [v for v in variables if v.get("Variable") not in
                  ("STUDYID", "RDOMAIN", "USUBJID", "IDVAR", "IDVARVAL",
                   "QNAM", "QLABEL", "QVAL", "QORIG", "QEVAL")]

    qual_list = []
    for q in qualifiers:
        qual_list.append(f"  QNAM='{q['Variable']}', QLABEL='{q.get('Label', q['Variable'])}'")

    return f"""You are a senior CDISC SDTM programmer. Generate a complete, production-quality SAS program
for the {domain} supplemental qualifier domain.

Parent domain: {parent}
RDOMAIN = '{parent}'
IDVAR = '{parent}SEQ'

SUPP structure (fixed columns):
{_var_table(structural)}

Qualifier variables to transpose into QNAM/QVAL rows:
{chr(10).join(qual_list)}

Requirements:
- Read source data from raw.{parent_lower} (the parent domain's source)
- Read sdtm.{parent} to get {parent}SEQ values for IDVARVAL
- For each qualifier variable, create one row per subject per parent record:
  - STUDYID = study constant
  - RDOMAIN = '{parent}'
  - USUBJID = from parent
  - IDVAR = '{parent}SEQ'
  - IDVARVAL = put({parent}SEQ, best.)
  - QNAM = variable name (e.g. '{qualifiers[0]["Variable"] if qualifiers else "QVAR"}')
  - QLABEL = variable label
  - QVAL = the actual data value (always character)
  - QORIG = 'CRF'
  - QEVAL = '' (blank unless evaluator-dependent)
- Stack all qualifier rows together using SET statements or PROC TRANSPOSE
- Only output rows where QVAL is non-missing
- Sort by STUDYID RDOMAIN USUBJID IDVAR IDVARVAL QNAM
- Output to sdtm.{domain}
- Include variable labels in a LABEL statement
- Use /*-- BEGIN {domain} --*/ and /*-- END {domain} --*/ markers

Generate the complete SAS program. No explanations, just the code."""


def _build_supp_prompt_r(domain, variables):
    var_table = _var_table(variables)
    parent = domain.replace("SUPP", "")
    parent_lower = parent.lower()

    structural = [v for v in variables if v.get("Variable") in
                  ("STUDYID", "RDOMAIN", "USUBJID", "IDVAR", "IDVARVAL",
                   "QNAM", "QLABEL", "QVAL", "QORIG", "QEVAL")]
    qualifiers = [v for v in variables if v.get("Variable") not in
                  ("STUDYID", "RDOMAIN", "USUBJID", "IDVAR", "IDVARVAL",
                   "QNAM", "QLABEL", "QVAL", "QORIG", "QEVAL")]

    qual_list = []
    for q in qualifiers:
        qual_list.append(f"  QNAM='{q['Variable']}', QLABEL='{q.get('Label', q['Variable'])}'")

    return f"""You are a senior CDISC SDTM programmer working in R (tidyverse). Generate a complete,
production-quality R script for the {domain} supplemental qualifier domain.

Parent domain: {parent}
RDOMAIN = "{parent}"
IDVAR = "{parent}SEQ"

SUPP structure (fixed columns):
{_var_table(structural)}

Qualifier variables to transpose into QNAM/QVAL rows:
{chr(10).join(qual_list)}

Requirements:
- library(dplyr), library(tidyr) at the top; read source data frame raw_{parent_lower} and the
  already-built {parent_lower} data frame (for {parent}SEQ values), both already in the R session
- For each qualifier variable, create one row per subject per parent record:
  - STUDYID = study constant
  - RDOMAIN = "{parent}"
  - USUBJID = from parent
  - IDVAR = "{parent}SEQ"
  - IDVARVAL = as.character({parent}SEQ)
  - QNAM = variable name (e.g. "{qualifiers[0]["Variable"] if qualifiers else "QVAR"}")
  - QLABEL = variable label
  - QVAL = the actual data value (always character)
  - QORIG = "CRF"
  - QEVAL = NA_character_ (unless evaluator-dependent)
- Stack all qualifier rows together — pivot_longer() from wide qualifier columns to QNAM/QVAL rows,
  or bind_rows() of one data frame per qualifier
- Only keep rows where QVAL is non-missing (filter(!is.na(QVAL)))
- Assign the final result to a data frame named {domain.lower()}
- Sort with arrange(STUDYID, RDOMAIN, USUBJID, IDVAR, IDVARVAL, QNAM)
- R has no length/label/format statements — do NOT emit them
- Use # -- BEGIN {domain} -- # and # -- END {domain} -- # markers

Generate the complete R script. No explanations, just the code, no markdown fences."""


# ── Build prompt router ─────────────────────────────────────────────

def _macro_hint_block(domain, variables, language):
    """Company macros (scope=sdtm) whose suffix pattern matches a variable
    in this domain, formatted as a prompt hint. SAS only — the catalog has
    no R macros, same as ADaM's R path. Unlike ADaM's per-variable exact
    match (which can skip the Writer entirely), SDTM generates a whole
    domain per prompt, so a matching macro is offered as a suggestion for
    the Writer to use where it fits, not a forced substitution — the Writer
    still decides whether/how to invoke it in context."""
    if language == "r":
        return ""
    var_names = [v["Variable"] for v in variables if v.get("Variable")]
    hints = find_sdtm_macros(var_names, _macro_catalog)
    if not hints:
        return ""
    lines = ["\n\nValidated company macros available for this domain — use them where "
            "they fit instead of writing the equivalent logic from scratch:"]
    for h in hints:
        lines.append(f"- {h['call']}\n  ({h['purpose']})")
    return "\n".join(lines)


def build_domain_prompt(domain, variables, language="sas", use_macros=True):
    """Route to the correct prompt builder based on domain class, then
    append any relevant company macro hints (see _macro_hint_block)."""
    dclass = get_domain_class(domain)

    if dclass == "DM":
        prompt = build_dm_prompt(domain, variables, language)
    elif dclass == "Events":
        prompt = build_events_prompt(domain, variables, language)
    elif dclass == "Interventions":
        prompt = build_interventions_prompt(domain, variables, language)
    elif dclass == "Findings":
        prompt = build_findings_prompt(domain, variables, language)
    elif dclass == "Findings About Events":
        prompt = build_fae_prompt(domain, variables, language)
    elif dclass == "SUPP":
        prompt = build_supp_prompt(domain, variables, language)
    else:
        prompt = build_events_prompt(domain, variables, language)

    if use_macros:
        prompt += _macro_hint_block(domain, variables, language)
    return prompt


# ── Three-agent pipeline ────────────────────────────────────────────

def generate_domain_program(domain, variables, use_api=True, language="sas", use_macros=True):
    """
    Generate a complete SAS or R program for one SDTM domain
    using the three-agent pipeline: Writer -> Improver -> Reviewer.
    """
    dclass = get_domain_class(domain)
    prompt = build_domain_prompt(domain, variables, language, use_macros)
    lang_name = "R" if language == "r" else "SAS"

    print(f"\n  [{domain}] Generating {lang_name} program ({dclass}, {len(variables)} variables)")

    # Step 1: Writer
    print(f"    Writer: ", end="", flush=True)
    if use_api:
        draft = generate_api(prompt)
        writer_model = API_MODEL
        print(f"API ({API_MODEL})")
    else:
        draft = generate_local(prompt)
        writer_model = LOCAL_MODEL
        print(f"Local ({LOCAL_MODEL})")

    if not draft or len(draft.strip()) < 50:
        print(f"    WARNING: Writer produced empty/short output for {domain}")
        return None, writer_model

    # Step 2: Improver
    # Step 2: Improver (direct API call, not the ADaM variable-level improver)
    if use_api:
        print(f"    Improver: API ({API_MODEL})")
        var_names = [v["Variable"] for v in variables if v.get("Variable")]
        if language == "r":
            improve_prompt = f"""You are a principal R programmer (tidyverse) with 15+ years of CDISC SDTM experience.
Review and improve this R script for the {domain} domain.

The script must create all these variables: {', '.join(var_names)}

Fix any issues:
- Missing or incorrect variable derivations
- Wrong types (character vs numeric, typed NA_character_/NA_real_)
- Missing select() at the end restricting to the spec variables
- Missing or incorrect sort order (arrange())
- Hardcoded values that should be derived
- Non-standard date handling (must parse ISO 8601 --DTC strings via as.Date(substr(x,1,10)))
- Do NOT flag valid R for not looking like SAS (no run;, no length/label/format, no semicolons)

Return ONLY the improved R code, no explanations, no markdown fences.""" + "\n\n" + draft
        else:
            improve_prompt = f"""You are a principal SAS programmer with 15+ years of CDISC SDTM experience.
Review and improve this SAS program for the {domain} domain.

The program must create all these variables: {', '.join(var_names)}

Fix any issues:
- Missing or incorrect variable derivations
- Wrong lengths, types, or formats
- Missing LABEL statements
- Missing or incorrect sort order
- Hardcoded values that should be derived
- Non-standard date handling (must be ISO 8601 character)

Return ONLY the improved SAS code, no explanations.""" + "\n\n" + draft

        try:
            improved = generate_api(improve_prompt)
            if improved and len(improved.strip()) > 50:
                draft = improved
            else:
                print(f"    Improver returned empty, keeping Writer draft")
        except Exception as e:
            print(f"    Improver failed: {e}, keeping Writer draft")

    # Step 3: Reviewer
    if use_api:
        print(f"    Reviewer: API ({API_MODEL})")
        review = review_sas(draft)
        if review:
            print(f"    Review: {review[:100]}...")
    else:
        review = None

    return draft, writer_model


# ── Program assembly ────────────────────────────────────────────────

def assemble_program(domain, code, variables, language="sas"):
    """
    Wrap the generated code in a standard SDTM program structure:
    header comment, libname/library, code, verification footer.
    """
    domain_lower = domain.lower()
    dclass = get_domain_class(domain)
    n_vars = len(variables)

    # Build variable list for the header
    var_names = [v["Variable"] for v in variables if v.get("Variable")]

    if language == "r":
        ext = "R"
        header = f"""# ********************************************************************
# Program:    {domain_lower}.R
# Domain:     {domain} ({dclass})
# Purpose:    Create SDTM {domain} domain data frame
# Variables:  {n_vars}
# Generated:  SpecGen Phase 5c - SDTM Program Generation (target = r)
#
# Output:     {domain_lower} data frame ({domain} domain dataset)
#
# Variables:  {', '.join(var_names[:8])}
#             {'...' if len(var_names) > 8 else ''}
# ********************************************************************
"""
        footer = f"""
# -- Verification -- #
# glimpse({domain_lower})
# table({domain_lower}$DOMAIN)

# End of {domain_lower}.R
"""
    else:
        ext = "sas"
        header = f"""/*******************************************************************************
* Program:    {domain_lower}.sas
* Domain:     {domain} ({dclass})
* Purpose:    Create SDTM {domain} domain dataset
* Variables:  {n_vars}
* Generated:  SpecGen Phase 5c - SDTM Program Generation
*
* Input:      raw.{domain_lower} (source CRF data)
* Output:     sdtm.{domain_lower} ({domain} domain dataset)
*
* Variables:  {', '.join(var_names[:8])}
*             {'...' if len(var_names) > 8 else ''}
*******************************************************************************/

/*-- Library assignments --*/
libname raw  'C:\\sas_data\\raw'  access=readonly;
libname sdtm 'C:\\sas_data\\sdtm';
"""
        footer = f"""
/*-- Final sort and output verification --*/
proc sort data=sdtm.{domain_lower};
  by STUDYID USUBJID;
run;

proc contents data=sdtm.{domain_lower} varnum;
run;

proc freq data=sdtm.{domain_lower};
  tables DOMAIN / nocum nopercent;
run;

/* End of {domain_lower}.sas */
"""

    # Clean up the generated code
    clean_code = code.strip()
    # Remove any markdown fences
    if clean_code.startswith("```"):
        clean_code = clean_code.split("\n", 1)[1]
    if clean_code.endswith("```"):
        clean_code = clean_code.rsplit("```", 1)[0]
    clean_code = clean_code.strip()

    return header + "\n" + clean_code + "\n" + footer


# ── Main entry points ───────────────────────────────────────────────

def generate_single_domain(xlsx_path, domain, output_dir, use_api=True, force=False, language="sas",
                           use_macros=True):
    """Generate the SAS or R program for one domain.

    By default, skips generation if output_dir/<domain>.<ext> already exists —
    each Writer/Improver run is a fresh, non-deterministic draft, so
    regenerating on every call would silently overwrite any hand-QC fix
    applied to a prior draft (see ROADMAP.md Phase 10 piece 3). Pass
    force=True to regenerate anyway (e.g. after a spec change).
    """
    domain_lower = domain.lower()
    ext = "R" if language == "r" else "sas"
    output_file = os.path.join(output_dir, f"{domain_lower}.{ext}")
    if not force and os.path.exists(output_file):
        print(f"  [{domain}] Skipping — {output_file} already exists (use --force to regenerate)")
        return output_file

    variables = read_domain_spec(xlsx_path, domain)
    if not variables:
        print(f"  No variables found for domain {domain}")
        return None

    code, writer_model = generate_domain_program(domain, variables, use_api=use_api, language=language,
                                                 use_macros=use_macros)
    if not code:
        print(f"  Failed to generate code for {domain}")
        return None

    program = assemble_program(domain, code, variables, language=language)

    # Write to file
    os.makedirs(output_dir, exist_ok=True)
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(program)

    print(f"    Written to {output_file} ({len(program)} chars)")

    # Log the run
    try:
        with _LOG_LOCK:
            log_run(
                spec_file=xlsx_path,
                mode="sdtm_generate",
                writer_model=writer_model,
                improver_model=API_MODEL if use_api else "none",
                reviewer_model=API_MODEL if use_api else "none",
                n_vars=len(variables),
                output_file=output_file,
            )
    except Exception as e:
        print(f"    Warning: could not log run: {e}")

    return output_file


SAS_FOOTER_MARKER = "/*-- Final sort and output verification --*/"
R_FOOTER_MARKER = "# -- Verification -- #"


def append_supp_domain(xlsx_path, supp_domain, output_dir, use_api=True, force=False, language="sas",
                       use_macros=True):
    """Generate a SUPP-- domain and append its code into its PARENT domain's
    output file (e.g. SUPPAE lives inside ae.sas/ae.R) instead of writing a
    separate program. SUPP-- is a supplemental-qualifier view of the same
    dataset, not an independent domain — most SOPs build it in the same
    program as its parent, since it shares the same source pull.

    Skips (like generate_single_domain) if the parent file already has this
    SUPP domain's BEGIN marker, unless force=True, in which case the old
    block is cut out and replaced with a fresh one.
    """
    ext = "R" if language == "r" else "sas"
    footer_marker = R_FOOTER_MARKER if language == "r" else SAS_FOOTER_MARKER
    parent = supp_domain.replace("SUPP", "")
    parent_file = os.path.join(output_dir, f"{parent.lower()}.{ext}")
    if not os.path.exists(parent_file):
        print(f"  [{supp_domain}] Parent domain file {parent_file} not found — skipping")
        return None

    with open(parent_file, encoding="utf-8") as f:
        parent_code = f.read()

    if language == "r":
        begin_marker = f"# -- BEGIN {supp_domain} -- #"
        end_marker = f"# -- END {supp_domain} -- #"
    else:
        begin_marker = f"/*-- BEGIN {supp_domain} --*/"
        end_marker = f"/*-- END {supp_domain} --*/"
    already_present = begin_marker in parent_code

    if already_present and not force:
        print(f"  [{supp_domain}] Skipping — already present in {parent_file} (use --force to regenerate)")
        return parent_file

    variables = read_domain_spec(xlsx_path, supp_domain)
    if not variables:
        print(f"  No variables found for domain {supp_domain}")
        return None

    code, writer_model = generate_domain_program(supp_domain, variables, use_api=use_api, language=language,
                                                 use_macros=use_macros)
    if not code:
        print(f"  Failed to generate code for {supp_domain}")
        return None

    clean_code = code.strip()
    if clean_code.startswith("```"):
        clean_code = clean_code.split("\n", 1)[1]
    if clean_code.endswith("```"):
        clean_code = clean_code.rsplit("```", 1)[0]
    clean_code = clean_code.strip()
    if language == "r":
        # The SUPP prompt asks for library(dplyr)/library(tidyr) since each
        # domain's generation is standalone, but here it's being appended
        # into a parent file that already loaded them in its own header —
        # strip redundant library() lines rather than duplicate them.
        clean_code = "\n".join(
            ln for ln in clean_code.splitlines() if not ln.strip().startswith("library(")
        ).strip()
    supp_block = f"\n{clean_code}\n"

    if already_present:  # force=True got us here — cut out the stale block first
        pattern = re.compile(re.escape(begin_marker) + r".*?" + re.escape(end_marker), re.DOTALL)
        parent_code = pattern.sub(clean_code, parent_code, count=1)
    elif footer_marker in parent_code:
        parent_code = parent_code.replace(footer_marker, supp_block + "\n" + footer_marker, 1)
    else:
        parent_code = parent_code + supp_block

    with open(parent_file, "w", encoding="utf-8") as f:
        f.write(parent_code)

    print(f"    Appended {supp_domain} into {parent_file} ({len(clean_code)} chars)")

    try:
        with _LOG_LOCK:
            log_run(
                spec_file=xlsx_path,
                mode="sdtm_generate",
                writer_model=writer_model,
                improver_model=API_MODEL if use_api else "none",
                reviewer_model=API_MODEL if use_api else "none",
                n_vars=len(variables),
                output_file=parent_file,
            )
    except Exception as e:
        print(f"    Warning: could not log run: {e}")

    return parent_file


def _run_concurrently(fn, domain_list, max_workers, xlsx_path, output_dir, use_api, force, language,
                      use_macros=True):
    """Run fn(xlsx_path, domain, output_dir, use_api, force=force, language=language,
    use_macros=use_macros) for every domain in domain_list at once (up to
    max_workers in flight),
    instead of one at a time. Each domain's own Writer->Improver->Reviewer
    calls are the slow part (~60-100s observed, 3 sequential API round-trips
    generating/reviewing a full program) — the domains themselves don't
    depend on each other (SUPP domains are only run after ALL standard
    domains finish, once every possible parent file already exists), so
    there's no reason to wait for one to finish before starting the next.
    Returns (results dict, failed list) — one failing domain doesn't stop
    the others.
    """
    results = {}
    failed = []
    if not domain_list:
        return results, failed

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_domain = {
            executor.submit(fn, xlsx_path, domain, output_dir, use_api, force=force, language=language,
                            use_macros=use_macros): domain
            for domain in domain_list
        }
        for future in as_completed(future_to_domain):
            domain = future_to_domain[future]
            try:
                result = future.result()
            except Exception as e:
                print(f"  [{domain}] FAILED: {e}")
                result = None
            if result:
                results[domain] = result
            else:
                failed.append(domain)

    return results, failed


def generate_all_domains(xlsx_path, output_dir, use_api=True, domains=None, force=False,
                         language="sas", max_workers=5, use_macros=True):
    """
    Generate SAS or R programs for all domains in the spec.
    Processes standard domains first, then SUPP domains
    (SUPP needs parent domain to exist first) — within each phase, domains
    run CONCURRENTLY (up to max_workers at once) since they don't depend on
    each other; only the two phases themselves are sequential.

    Domains whose output file already exists are skipped unless force=True
    (see generate_single_domain) — repeated runs (e.g. every "Generate" click
    in the web app) are idempotent by default instead of overwriting hand-QC
    fixes with a fresh non-deterministic draft each time.
    """
    all_domains = list_domains(xlsx_path)

    if domains:
        all_domains = [d for d in all_domains if d in domains]

    std_domains = [d for d in all_domains if not d.startswith("SUPP")]
    supp_domains = [d for d in all_domains if d.startswith("SUPP")]

    print(f"SpecGen Phase 5c: SDTM Program Generation")
    print(f"  Spec: {xlsx_path}")
    print(f"  Output: {output_dir}")
    print(f"  Language: {'R' if language == 'r' else 'SAS'}")
    print(f"  Standard domains ({len(std_domains)}): {', '.join(std_domains)}")
    print(f"  SUPP domains ({len(supp_domains)}): {', '.join(supp_domains)}")
    print(f"  Mode: {'API' if use_api else 'Offline'}{' (force regenerate)' if force else ''}")
    print(f"  Concurrency: up to {max_workers} domains at once")

    # Standard domains first, concurrently — SUPP domains need every parent
    # file to exist before this phase starts, so it can't overlap the next
    results, failed = _run_concurrently(
        generate_single_domain, std_domains, max_workers,
        xlsx_path, output_dir, use_api, force=force, language=language, use_macros=use_macros,
    )

    # SUPP domains after (they reference parent domain) — appended into the
    # parent domain's own output file, not written as a separate program.
    # Each SUPP domain has its own distinct parent (1:1 by CDISC convention),
    # so running them concurrently with each other is safe too.
    supp_results, supp_failed = _run_concurrently(
        append_supp_domain, supp_domains, max_workers,
        xlsx_path, output_dir, use_api, force=force, language=language, use_macros=use_macros,
    )
    results.update(supp_results)
    failed.extend(supp_failed)

    # Summary
    print(f"\n{'='*60}")
    print(f"SDTM Generation Summary")
    print(f"{'='*60}")
    print(f"  Generated: {len(results)} programs")
    for domain, path in sorted(results.items()):
        dclass = get_domain_class(domain)
        print(f"    {domain:<10} [{dclass}] -> {path}")
    if failed:
        print(f"  Failed: {len(failed)} - {', '.join(failed)}")
    print(f"{'='*60}")

    return results


# ── CLI entry point ─────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate SDTM SAS programs from a draft SDTM specification.")
    parser.add_argument("spec", help="Path to sdtm_spec_draft.xlsx")
    parser.add_argument("--output", "-o", default="sdtm_programs",
                        help="Output directory (default: sdtm_programs/)")
    parser.add_argument("--domain", "-d",
                        help="Generate for a single domain (e.g. DM, AE, VS)")
    parser.add_argument("--offline", action="store_true",
                        help="Use local Ollama only, no API calls")
    parser.add_argument("--force", action="store_true",
                        help="Regenerate domains even if their output file already "
                             "exists (default: skip existing files, so hand-QC fixes "
                             "aren't overwritten by a fresh draft)")
    parser.add_argument("--lang", choices=["sas", "r"], default="sas",
                        help="Output language (default: sas)")
    parser.add_argument("--workers", "-w", type=int, default=5,
                        help="Max domains to generate concurrently (default: 5). "
                             "Domains are independent API-bound work, so running "
                             "several at once is much faster than one at a time; "
                             "raise/lower to match your API rate limit.")
    parser.add_argument("--no-macros", action="store_true",
                        help="Don't offer the validated company macro catalog as a "
                             "hint to the Writer (default: offered where a domain's "
                             "variables match one, e.g. --DTC/--SEQ/SUPP qualifiers)")

    args = parser.parse_args()

    if args.domain:
        domains = [d.strip().upper() for d in args.domain.split(",")]
        generate_all_domains(args.spec, args.output,
                             use_api=not args.offline, domains=domains, force=args.force,
                             language=args.lang, max_workers=args.workers,
                             use_macros=not args.no_macros)
    else:
        generate_all_domains(args.spec, args.output,
                             use_api=not args.offline, force=args.force, language=args.lang,
                             max_workers=args.workers, use_macros=not args.no_macros)
