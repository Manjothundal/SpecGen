"""
build_sample_shell_vs.py - sample VS shift-table mock shell.

The third shell claude_code_instructions.md asks for, alongside the
existing demographics (build_sample_shell.py) and AE summary
(build_sample_shell_ae.py) shells — those two already exist and are
reused unchanged by adam_bds_assembler.py, so this only builds the new
one: a shift table needs to know which PARAMCDs are in scope and each
one's normal range, to derive a Low/Normal/High shift category.

Two sheets, same Shell_Meta/Shell_Rows convention as every other shell in
this app (tlf_assembler.py's _read_shell / adam_bds_assembler.py's
read_shell both read this unchanged):
  Shell_Meta  — titles, population, source dataset (ADVS), baseline visit,
                column var, footnotes
  Shell_Rows  — order, paramcd, param, low_cutoff, high_cutoff — one row
                per vital sign parameter the shift table covers
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def build_vs_shift_shell(path="sample_shell_vs_shift.xlsx"):
    wb = openpyxl.Workbook()

    meta = wb.active
    meta.title = "Shell_Meta"
    for r in [
        ["Field", "Value"],
        ["table_id", "14.4.1"],
        ["title1", "Table 14.4.1"],
        ["title2", "Shift Table for Vital Signs: Baseline vs. Worst On-Treatment Value"],
        ["title3", "Safety Population"],
        ["population", "SAFFL"],
        ["source_dataset", "ADVS"],
        ["baseline_visit", "BASELINE"],
        ["column_var", "TRT01A"],
        ["add_total_column", "YES"],
        ["footnote1", "N = number of subjects in the safety population with a "
                      "non-missing value at both baseline and the visit shown."],
        ["footnote2", "Shift categories: Low / Normal / High, per the range shown for each parameter."],
    ]:
        meta.append(r)

    rows = wb.create_sheet("Shell_Rows")
    rows.append(["order", "paramcd", "param", "low_cutoff", "high_cutoff"])
    for r in [
        [1, "SYSBP", "Systolic Blood Pressure (mmHg)", 90, 140],
        [2, "DIABP", "Diastolic Blood Pressure (mmHg)", 60, 90],
        [3, "PULSE", "Pulse Rate (beats/min)", 60, 100],
    ]:
        rows.append(r)

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    for ci in range(1, 3):
        cell = meta.cell(row=1, column=ci)
        cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, header_align, thin_border
    meta.column_dimensions["A"].width = 18
    meta.column_dimensions["B"].width = 55

    for ci, w in enumerate([8, 12, 32, 12, 12], 1):
        cell = rows.cell(row=1, column=ci)
        cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, header_align, thin_border
        rows.column_dimensions[get_column_letter(ci)].width = w

    wb.save(path)
    print(f"Wrote {path}")
    print("  Shell_Meta: source_dataset=ADVS, population=SAFFL, baseline_visit=BASELINE")
    print("  Shell_Rows: 3 PARAMCDs (SYSBP, DIABP, PULSE) with normal-range cutoffs")


if __name__ == "__main__":
    build_vs_shift_shell()
