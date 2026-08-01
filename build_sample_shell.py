"""
build_sample_shell.py — generate a sample TLF mock shell (Table 14.1.1
Demographics) as a structured Excel file the TLF generator can read.

A mock shell is the empty table layout: title, column headers (one per
treatment arm + Total), the row structure (which variables and which
statistics), and footnotes. No numbers — it's the template.

We use Excel (not a Word/RTF shell) because it's machine-readable: the
generator reads rows/columns directly instead of parsing a formatted doc.
A real project's shells are usually Word/RTF; parsing those is a later
enhancement (mirrors how aCRF parsing handles real PDFs).
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def build_demographics_shell(path="sample_shell_demographics.xlsx"):
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Shell metadata (titles, population, footnotes) ----
    meta = wb.active
    meta.title = "Shell_Meta"
    meta_rows = [
        ["Field", "Value"],
        ["table_id", "14.1.1"],
        ["title1", "Table 14.1.1"],
        ["title2", "Summary of Demographic and Baseline Characteristics"],
        ["title3", "Safety Population"],
        ["population", "SAFFL"],          # which ADSL flag defines the population
        ["source_dataset", "ADSL"],
        ["column_var", "TRT01A"],         # what defines the treatment columns
        ["add_total_column", "YES"],
        ["footnote1", "N = number of subjects in the safety population."],
        ["footnote2", "Percentages are based on N within each treatment group."],
    ]
    for r in meta_rows:
        meta.append(r)

    # ---- Sheet 2: Row structure (what to show, and how) ----
    rows = wb.create_sheet("Shell_Rows")
    header = ["order", "adam_var", "label", "stat_type", "decimals"]
    rows.append(header)
    # stat_type drives which summary the generator computes:
    #   contn  = n, mean, sd, median, min, max  (continuous)
    #   catn   = n (%) per category             (categorical)
    row_defs = [
        [1, "AGE",    "Age (years)",        "contn", 1],
        [2, "AGEGR1", "Age Group, n (%)",   "catn",  0],
        [3, "SEX",    "Sex, n (%)",         "catn",  0],
        [4, "RACE",   "Race, n (%)",        "catn",  0],
        [5, "BMIBL",  "Baseline BMI (kg/m2)","contn", 1],
    ]
    for r in row_defs:
        rows.append(r)

    # light styling so it reads like a shell
    for ws in (meta, rows):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    wb.save(path)
    print(f"Wrote {path}")
    print("  Sheet 'Shell_Meta': titles, population (SAFFL), column var (TRT01A), footnotes")
    print("  Sheet 'Shell_Rows': 5 rows — AGE, AGEGR1, SEX, RACE, BMIBL (contn/catn)")


if __name__ == "__main__":
    build_demographics_shell()
