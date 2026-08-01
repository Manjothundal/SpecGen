"""
build_sample_shell_ae.py — sample AE summary shell (Table 14.3.1).

An AE summary counts DISTINCT SUBJECTS with >=1 event meeting a condition —
not rows, since one subject can have many AEs. Rows are condition-based
(any TEAE / serious / related / by max severity), each a filter over ADAE.

Two sheets:
  Shell_Meta  — titles, population (ADSL SAFFL), numerator source (ADAE),
                treatment column var, denominator source, footnotes
  Shell_Rows  — order, row_label, condition (a filter expression over ADAE),
                indent (for sub-rows under a heading)

The 'condition' is a plain expression the generator turns into SAS/R filter
logic. Empty condition = the "any TEAE" baseline (just TRTEMFL='Y').
"""

import openpyxl
from openpyxl.styles import Font

def build_ae_shell(path="sample_shell_ae.xlsx"):
    wb = openpyxl.Workbook()

    meta = wb.active
    meta.title = "Shell_Meta"
    for r in [
        ["Field", "Value"],
        ["table_id", "14.3.1"],
        ["title1", "Table 14.3.1"],
        ["title2", "Overall Summary of Treatment-Emergent Adverse Events"],
        ["title3", "Safety Population"],
        ["population", "SAFFL"],           # denominator flag in ADSL
        ["denom_dataset", "ADSL"],         # where big-N comes from
        ["numer_dataset", "ADAE"],         # where events come from
        ["teae_flag", "TRTEMFL"],          # treatment-emergent flag
        ["column_var", "TRT01A"],
        ["add_total_column", "YES"],
        ["footnote1", "TEAE = treatment-emergent adverse event."],
        ["footnote2", "A subject is counted once within each row, regardless of the number of events."],
        ["footnote3", "Percentages use the number of safety-population subjects per arm as denominator."],
    ]:
        meta.append(r)

    rows = wb.create_sheet("Shell_Rows")
    rows.append(["order", "row_label", "condition", "indent"])
    # condition is an ADAE-level filter; blank = all TEAE (TRTEMFL='Y' only)
    for r in [
        [1, "Subjects with any TEAE", "", 0],
        [2, "Subjects with any serious TEAE", 'AESER = "Y"', 0],
        [3, "Subjects with any drug-related TEAE", 'AEREL = "Y"', 0],
        [4, "Subjects with any TEAE leading to discontinuation", 'AEACN = "DRUG WITHDRAWN"', 0],
        [5, "TEAE by maximum severity", "", 0],          # heading only
        [6, "Mild", 'AESEV = "MILD"', 1],
        [7, "Moderate", 'AESEV = "MODERATE"', 1],
        [8, "Severe", 'AESEV = "SEVERE"', 1],
    ]:
        rows.append(r)

    for ws in (meta, rows):
        for c in ws[1]:
            c.font = Font(bold=True)

    wb.save(path)
    print(f"Wrote {path}")
    print("  Shell_Meta: denom=ADSL/SAFFL, numer=ADAE, teae_flag=TRTEMFL, column_var=TRT01A")
    print("  Shell_Rows: 8 rows (any/serious/related/discon + severity Mild/Moderate/Severe)")


if __name__ == "__main__":
    build_ae_shell()
